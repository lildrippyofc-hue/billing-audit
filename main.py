import os
import secrets
import hashlib
import sqlite3
import json
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, List, Any, Dict
from zoneinfo import ZoneInfo
from io import BytesIO
from collections import defaultdict

import requests as req_lib
from requests import Session as ReqSession

from fastapi import FastAPI, HTTPException, Cookie, Response, Depends, UploadFile, File, Form
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# ── Paths & DB ────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent

# On Railway set DATA_DIR to your mounted volume path so data survives deploys.
# Locally it defaults to the project folder (same as before).
DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR)))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "audit.db"
EXPORTS_DIR = DATA_DIR / "shift_exports"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
HISTORICAL_DATA_PATH = BASE_DIR / "historical-data.js"

try:
    DMS_BUSINESS_TZ = ZoneInfo("America/Chicago")
except Exception:
    DMS_BUSINESS_TZ = timezone(timedelta(hours=int(os.environ.get("DMS_BUSINESS_UTC_OFFSET_HOURS", "-5"))))
DMS_NEXT_BUSINESS_DATE_HOUR = int(os.environ.get("DMS_NEXT_BUSINESS_DATE_HOUR", "19"))

# ── Auth config ───────────────────────────────────────────────────────────────

def _hash(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

# To change the password set the APP_PASSWORD environment variable on Railway.
# Default password for local dev only — override it in production!
_APP_PASSWORD = os.environ.get("APP_PASSWORD", "N3747P9R")

# Role access system:
# admin   -> full access
# oks     -> ALDIOKS live portal and truck report
# minnesota -> Minnesota portal and truck report
# manager -> operations, audits, reporting
# lead    -> live board and lead tools
# clerk   -> live board, DMS stamp, clerk support
# client  -> clean read-only reporting
# guest   -> basic read-only reporting
_USERS: Dict[str, str] = {
    "james":   _hash(_APP_PASSWORD),
    "aldioks": _hash(os.environ.get("ALDIOKS_PASSWORD", os.environ.get("APP_PASSWORD", "N3747P9R"))),
    "minnesota": _hash(os.environ.get("MINNESOTA_PASSWORD", "mn1")),
    "dean":    _hash(_APP_PASSWORD),
    "manager": _hash(os.environ.get("MANAGER_PASSWORD", "manager1")),
    "teamlead": _hash(os.environ.get("TEAMLEAD_PASSWORD", "lead1")),
    "clerk":   _hash(os.environ.get("CLERK_PASSWORD", "clerk1")),
    "client":  _hash(os.environ.get("CLIENT_PASSWORD", "client1")),
    "work":    _hash(os.environ.get("WORK_PASSWORD", "work1")),
    "guest":   _hash(os.environ.get("GUEST_PASSWORD", "guest1")),
}

# Role lookup
_ROLES: Dict[str, str] = {
    "james":   "admin",
    "aldioks": "oks",
    "minnesota": "minnesota",
    "dean":    "admin",
    "manager": "manager",
    "teamlead": "teamlead",
    "clerk":   "clerk",
    "client":  "client",
    "work":    "work",
    "guest":   "guest",
}

# In-memory session store (fine for a single-process server)
_sessions: Dict[str, str] = {}

# ── App setup ─────────────────────────────────────────────────────────────────

app = FastAPI(title="Billing Audit API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)


# ── DB helpers ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS reports (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            filename    TEXT    NOT NULL,
            uploaded_at TEXT    NOT NULL,
            row_count   INTEGER NOT NULL DEFAULT 0,
            headers     TEXT    NOT NULL,
            rows        TEXT    NOT NULL,
            selectors   TEXT    NOT NULL
        );

        CREATE TABLE IF NOT EXISTS audit_decisions (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            truck_key    TEXT    NOT NULL,
            decision     TEXT    NOT NULL,
            report_id    INTEGER,
            po_keys      TEXT    DEFAULT '[]',
            supplier_key TEXT,
            decided_at   TEXT    NOT NULL,
            UNIQUE(truck_key),
            FOREIGN KEY (report_id) REFERENCES reports(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS visits (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            username   TEXT    NOT NULL,
            visited_at TEXT    NOT NULL
        );

        CREATE TABLE IF NOT EXISTS vendor_unload_times (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor      TEXT    NOT NULL,
            dock_min    INTEGER NOT NULL,
            shift_date  TEXT    NOT NULL,
            source      TEXT    NOT NULL DEFAULT 'Manual',
            truck_ref   TEXT,
            recorded_at TEXT    NOT NULL
        );

        CREATE TABLE IF NOT EXISTS billing_audit_runs (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            site               TEXT    NOT NULL,
            week_start         TEXT    NOT NULL,
            run_at             TEXT    NOT NULL DEFAULT CURRENT_TIMESTAMP,
            created_by_user    TEXT,
            ops_filename       TEXT,
            ops_sha256         TEXT,
            result_sha256      TEXT,
            result_json        TEXT    NOT NULL,
            is_baseline        INTEGER NOT NULL DEFAULT 0,
            deleted_at         TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_billing_audit_runs_key
            ON billing_audit_runs(site, week_start, run_at);
    """)
    conn.commit()
    conn.close()


init_db()


# ── Pydantic models ───────────────────────────────────────────────────────────

class ReportIn(BaseModel):
    filename: str
    headers: List[str]
    rows: List[Dict[str, Any]]
    selectors: Dict[str, str]


class DecisionIn(BaseModel):
    truck_key: str
    decision: str
    report_id: Optional[int] = None
    po_keys: Optional[List[str]] = []
    supplier_key: Optional[str] = None


class LoginIn(BaseModel):
    username: str
    password: str


_dms_session_cache: Dict[str, Any] = {}
_dms_mn_session_cache: Dict[str, Any] = {}


def _load_dms_config() -> Dict[str, Any]:
    cfg_path = BASE_DIR / "dms_config.json"
    cfg: Dict[str, Any] = {}
    parse_error: str = ""
    if cfg_path.exists():
        try:
            raw = cfg_path.read_text(encoding="utf-8-sig")  # utf-8-sig strips BOM if present
            cfg = json.loads(raw)
        except Exception as e:
            parse_error = str(e)
            cfg = {}
    username = os.environ.get("DMS_USERNAME") or cfg.get("username") or ""
    password = os.environ.get("DMS_PASSWORD") or cfg.get("password") or ""
    base = (os.environ.get("DMS_BASE_URL") or cfg.get("base_url") or "https://dms.eclipseia.com").rstrip("/")
    # The DMS API is served on 443; :5055 is dead and only causes connect timeouts.
    base = base.replace(":5055", "")
    return {
        "username": username,
        "password": password,
        "base_url": base,
        "location_code": os.environ.get("DMS_LOCATION_CODE") or cfg.get("location_code") or "OLA",
        "location_name": os.environ.get("DMS_LOCATION_NAME") or cfg.get("location_name") or "ALDIOKS",
        "timeout": int(os.environ.get("DMS_TIMEOUT_SECONDS") or cfg.get("timeout_seconds") or 25),
        "_parse_error": parse_error,
        "_cfg_path": str(cfg_path),
    }


def _load_dms_mn_config() -> Dict[str, Any]:
    cfg_path = BASE_DIR / "dms_mn_config.json"
    cfg: Dict[str, Any] = {}
    parse_error: str = ""
    if cfg_path.exists():
        try:
            raw = cfg_path.read_text(encoding="utf-8-sig")
            cfg = json.loads(raw)
        except Exception as e:
            parse_error = str(e)
            cfg = {}
    username = os.environ.get("DMS_MN_USERNAME") or cfg.get("username") or ""
    password = os.environ.get("DMS_MN_PASSWORD") or cfg.get("password") or ""
    base = (os.environ.get("DMS_MN_BASE_URL") or cfg.get("base_url") or "https://dms.eclipseia.com").rstrip("/")
    base = base.replace(":5055", "")
    return {
        "username": username,
        "password": password,
        "base_url": base,
        "location_id": os.environ.get("DMS_MN_LOCATION_ID") or cfg.get("location_id") or "",
        "location_code": os.environ.get("DMS_MN_LOCATION_CODE") or cfg.get("location_code") or "",
        "location_name": os.environ.get("DMS_MN_LOCATION_NAME") or cfg.get("location_name") or "MINNESOTA",
        "timeout": int(os.environ.get("DMS_MN_TIMEOUT_SECONDS") or cfg.get("timeout_seconds") or 25),
        "_parse_error": parse_error,
        "_cfg_path": str(cfg_path),
    }


_EDGE_HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json, text/plain, */*",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",
}


def _dms_post(path: str, payload: Dict[str, Any], config: Dict[str, Any]) -> Any:
    url = f"{config['base_url']}/{path.lstrip('/')}"
    session = config.get("_session") or req_lib
    try:
        resp = session.post(url, json=payload, headers=_EDGE_HEADERS, timeout=config["timeout"])
        if not resp.ok:
            raise HTTPException(status_code=502, detail=f"DMS returned {resp.status_code} for {path}: {resp.text[:500]}")
        return resp.json() if resp.text.strip() else {}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Could not reach DMS ({path}): {exc}")


def _dms_json_request(path: str, payload: Dict[str, Any], config: Dict[str, Any]) -> Any:
    return _dms_post(path, payload, config)


def _dms_login_payloads(username: str, password: str) -> List[Dict[str, Any]]:
    return [
        {"un": username, "pw": password},
        {"username": username, "password": password},
        {"user": username, "password": password},
    ]


def _first_list(value: Any) -> List[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        # NOTE: DMS api/stamp/getStamps wraps the rows under "list".
        for key in ("list", "data", "rows", "loads", "stamps", "result", "results", "Table"):
            found = _first_list(value.get(key))
            if found:
                return found
    return []


def _find_dms_locations(login_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    candidates = []
    for key in ("locations", "locs", "location", "loc"):
        value = login_data.get(key)
        if isinstance(value, list):
            candidates.extend([x for x in value if isinstance(x, dict)])
        elif isinstance(value, dict):
            candidates.append(value)
    userinfo = login_data.get("userinfo")
    if isinstance(userinfo, dict):
        for key in ("locations", "locs", "location", "loc"):
            value = userinfo.get(key)
            if isinstance(value, list):
                candidates.extend([x for x in value if isinstance(x, dict)])
            elif isinstance(value, dict):
                candidates.append(value)
    return candidates


def _select_dms_location(locations: List[Dict[str, Any]], config: Dict[str, Any]) -> Dict[str, Any]:
    loc_id = str(config.get("location_id") or "").strip()
    code = str(config["location_code"]).upper()
    name = str(config["location_name"]).upper()
    if loc_id:
        for loc in locations:
            if str(loc.get("locid") or loc.get("id") or "").strip() == loc_id:
                return loc
    for loc in locations:
        text = " ".join(str(loc.get(k, "")) for k in ("cCode", "code", "name", "locName", "location", "locid")).upper()
        if code and code in text:
            return loc
    for loc in locations:
        text = " ".join(str(v) for v in loc.values()).upper()
        if name and name in text:
            return loc
    if locations:
        return locations[0]
    raise HTTPException(status_code=502, detail="DMS login worked, but no DMS location was returned.")


def _dms_location_matches_config(loc: Dict[str, Any], config: Dict[str, Any]) -> bool:
    loc_id = str(config.get("location_id") or "").strip()
    code = str(config.get("location_code") or "").strip().upper()
    name = str(config.get("location_name") or "").strip().upper()
    if loc_id and str(loc.get("locid") or loc.get("id") or "").strip() == loc_id:
        return True
    text = " ".join(str(loc.get(k, "")) for k in ("cCode", "code", "name", "locName", "location", "locid")).upper()
    all_text = " ".join(str(v) for v in loc.values()).upper()
    return bool((code and code in text) or (name and name in all_text))


def _dms_session_is_valid(session: Dict[str, Any]) -> bool:
    """A cached session is only useful if DMS actually returned a real
    location/user, not an empty/all-null body with a 2xx status (seen
    intermittently from the DMS login endpoint). Without this check, one bad
    login response gets cached forever and every later request silently fails
    until the process is restarted."""
    loc = session.get("loc")
    userinfo = session.get("userinfo")
    if not isinstance(loc, dict) or not loc.get("locid"):
        return False
    if not isinstance(userinfo, dict) or not (userinfo.get("locid") or userinfo.get("login")):
        return False
    return True


def _ensure_dms_session(force: bool = False) -> Dict[str, Any]:
    config = _load_dms_config()
    if (
        not force
        and _dms_session_cache.get("base_url") == config["base_url"]
        and _dms_session_is_valid(_dms_session_cache)
    ):
        return _dms_session_cache

    username = str(config["username"]).strip()
    password = str(config["password"]).strip()
    if not username or not password or "YOUR_" in username or "YOUR_" in password:
        parse_err = config.get("_parse_error", "")
        cfg_path  = config.get("_cfg_path", "dms_config.json")
        detail = (
            f"DMS credentials are not configured. "
            f"Config file: {cfg_path}. "
            + (f"JSON parse error: {parse_err}. " if parse_err else "File parsed OK but username/password missing. ")
            + "Fill in username and password in dms_config.json."
        )
        raise HTTPException(status_code=400, detail=detail)

    # Use a requests Session so cookies persist across the login + data calls
    dms_session = ReqSession()
    dms_session.headers.update(_EDGE_HEADERS)
    config["_session"] = dms_session

    # Seed cookies by loading the login page first
    try:
        dms_session.get(f"{config['base_url']}/login", timeout=config["timeout"])
    except Exception:
        pass

    last_error = None
    login_data: Dict[str, Any] = {}
    for payload in _dms_login_payloads(username, password):
        try:
            response = _dms_post("api/login/trylogin", payload, config)
            ui = response.get("userinfo") or {}
            if isinstance(response, dict) and ui.get("login"):
                login_data = response
                break
        except HTTPException as exc:
            last_error = exc
    if not login_data:
        if last_error:
            raise last_error
        raise HTTPException(status_code=502, detail="DMS login did not return session data. Credentials may be wrong.")

    # DMS returns selLoc directly on login — use it if present
    sel_loc = login_data.get("selLoc")
    if isinstance(sel_loc, dict) and sel_loc:
        loc = sel_loc
    else:
        locations = _find_dms_locations(login_data)
        if not locations:
            try:
                loc_response = _dms_json_request("api/location/getLocations", {"userinfo": login_data.get("userinfo") or login_data}, config)
                locations = [x for x in _first_list(loc_response) if isinstance(x, dict)]
            except HTTPException:
                locations = []
        loc = _select_dms_location(locations, config) if locations else sel_loc or {}

    session = {
        "base_url": config["base_url"],
        "userinfo": login_data.get("userinfo") or login_data.get("user") or login_data,
        "buck": login_data.get("buck") or login_data.get("bucket") or {},
        "loc": loc,
        "sel_loc": sel_loc or loc,
        "appts": login_data.get("appts") or [],
        "sel_appt": login_data.get("selAppt") or "",
        "config": config,
        "cached_at": datetime.now(timezone.utc).isoformat(),
    }
    _dms_session_cache.clear()
    _dms_session_cache.update(session)
    return session


def _ensure_dms_mn_session(force: bool = False) -> Dict[str, Any]:
    config = _load_dms_mn_config()
    cache_key = "|".join([
        config["base_url"],
        str(config.get("location_id") or ""),
        str(config.get("location_code") or ""),
        str(config.get("location_name") or ""),
    ])
    if (
        not force
        and _dms_mn_session_cache.get("cache_key") == cache_key
        and _dms_session_is_valid(_dms_mn_session_cache)
    ):
        return _dms_mn_session_cache

    username = str(config["username"]).strip()
    password = str(config["password"]).strip()
    if not username or not password or "YOUR_" in username or "YOUR_" in password:
        parse_err = config.get("_parse_error", "")
        cfg_path = config.get("_cfg_path", "dms_mn_config.json")
        detail = (
            f"Minnesota DMS credentials are not configured. "
            f"Config file: {cfg_path}. "
            + (f"JSON parse error: {parse_err}. " if parse_err else "File parsed OK but username/password missing. ")
            + "Set DMS_MN_USERNAME/DMS_MN_PASSWORD or fill in dms_mn_config.json."
        )
        raise HTTPException(status_code=400, detail=detail)

    dms_session = ReqSession()
    dms_session.headers.update(_EDGE_HEADERS)
    config["_session"] = dms_session

    try:
        dms_session.get(f"{config['base_url']}/login", timeout=config["timeout"])
    except Exception:
        pass

    last_error = None
    login_data: Dict[str, Any] = {}
    for payload in _dms_login_payloads(username, password):
        try:
            response = _dms_post("api/login/trylogin", payload, config)
            ui = response.get("userinfo") or {}
            if isinstance(response, dict) and ui.get("login"):
                login_data = response
                break
        except HTTPException as exc:
            last_error = exc
    if not login_data:
        if last_error:
            raise last_error
        raise HTTPException(status_code=502, detail="Minnesota DMS login did not return session data. Credentials may be wrong.")

    sel_loc = login_data.get("selLoc")
    loc: Dict[str, Any] = {}
    if isinstance(sel_loc, dict) and sel_loc and _dms_location_matches_config(sel_loc, config):
        loc = sel_loc
    else:
        locations = _find_dms_locations(login_data)
        if not locations:
            try:
                loc_response = _dms_json_request("api/location/getLocations", {"userinfo": login_data.get("userinfo") or login_data}, config)
                locations = [x for x in _first_list(loc_response) if isinstance(x, dict)]
            except HTTPException:
                locations = []
        loc = _select_dms_location(locations, config) if locations else (sel_loc if isinstance(sel_loc, dict) else {})

    session = {
        "base_url": config["base_url"],
        "userinfo": login_data.get("userinfo") or login_data.get("user") or login_data,
        "buck": login_data.get("buck") or login_data.get("bucket") or {},
        "loc": loc,
        "sel_loc": loc,
        "appts": login_data.get("appts") or [],
        "sel_appt": login_data.get("selAppt") or "",
        "config": config,
        "cache_key": cache_key,
        "cached_at": datetime.now(timezone.utc).isoformat(),
    }
    _dms_mn_session_cache.clear()
    _dms_mn_session_cache.update(session)
    return session


def _dms_business_date(date_text: Optional[str]) -> str:
    if date_text:
        for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
            try:
                dt = datetime.strptime(date_text, fmt)
                return f"{dt.month}/{dt.day}/{dt.year}"
            except ValueError:
                pass
    dt = datetime.now(DMS_BUSINESS_TZ)
    if dt.hour >= DMS_NEXT_BUSINESS_DATE_HOUR:
        dt = dt + timedelta(days=1)
    return f"{dt.month}/{dt.day}/{dt.year}"


def _parse_dms_time(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    # DMS stamp times look like "06/13/2026 03:46 AM UTC" (always UTC).
    # Strip a trailing UTC/GMT marker so the AM/PM formats below match; the
    # value is UTC either way and we tag it as such.
    text = text.replace("+00:00", "Z")
    low = text.lower()
    for suffix in (" utc", " gmt"):
        if low.endswith(suffix):
            text = text[: -len(suffix)].strip()
            break
    formats = [
        "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(text, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.isoformat()
        except ValueError:
            pass
    return text


def _dms_key(row: Dict[str, Any]) -> str:
    for key in ("rowid", "loadrowid", "ldrowid", "id", "loadid"):
        value = row.get(key)
        if value not in (None, ""):
            return f"id:{value}"
    for key in ("poNum", "po", "ponum", "trkNum", "trknum", "truck", "ref"):
        value = row.get(key)
        if value not in (None, ""):
            return f"{key}:{value}"
    return ""


def _normalize_portal_truck(load: Dict[str, Any], stamp: Dict[str, Any]) -> Dict[str, Any]:
    # Merge load (truck info: qty/desc/notes) with stamp (the timestamps), joined
    # by rowid. DMS field names confirmed from a live shift:
    #   drchk = driver check-in, drdoor = driver at door, clrkchk = clerk check-in,
    #   unstart = unload start, unfin = unload finish, recstart = receiving start,
    #   recfin = receiving finish, drleft = driver left, drstat = status.
    merged = {**load, **stamp}
    appointment = _parse_dms_time(
        merged.get("appt") or merged.get("apptDisplay") or merged.get("appointment") or merged.get("appointmentTime")
    )
    # Real driver/clerk check-in only — do NOT fall back to the appointment, or
    # every scheduled-but-not-arrived truck would look "checked in".
    check_in = _parse_dms_time(
        merged.get("drchk") or merged.get("driverCheckIn") or merged.get("driver_check_in") or merged.get("clrkchk")
    )
    driver_at_door = _parse_dms_time(
        merged.get("drdoor") or merged.get("driverAtDoor") or merged.get("driver_at_door")
    )
    unload_start = _parse_dms_time(
        merged.get("unstart") or merged.get("unloadStart") or merged.get("unload_start")
    )
    unload_finish = _parse_dms_time(
        merged.get("unfin") or merged.get("unloadFinish") or merged.get("unload_finish")
    )
    receiving_start = _parse_dms_time(
        merged.get("recstart") or merged.get("receivingStart") or merged.get("receiving_start")
    )
    receiving_finish = _parse_dms_time(
        merged.get("recfin") or merged.get("receivingFinish") or merged.get("receiving_finish")
    )
    driver_left = _parse_dms_time(merged.get("drleft") or merged.get("driverLeft"))
    ref = (
        merged.get("trkNum") or merged.get("trk") or merged.get("truck")
        or merged.get("cabNum") or merged.get("rowid") or merged.get("poNum") or ""
    )
    return {
        "id": f"dms-{_dms_key(merged) or ref}",
        "source": "DMS",
        "rowid": merged.get("rowid"),
        "ref": str(ref or "").strip(),
        "door": str(merged.get("doorNum") or merged.get("door") or "").strip(),
        "supplier": str(merged.get("sup") or merged.get("supplier") or merged.get("vendor") or "").strip(),
        "carrier": str(merged.get("trnum") or merged.get("carr") or "").strip(),
        "po": str(merged.get("poNum") or merged.get("po") or "").strip(),
        "area": str(merged.get("area") or "").strip(),
        "comments": str(merged.get("comments") or merged.get("notes") or "").strip(),
        "appointmentIso": appointment,
        "checkInIso": check_in,
        "driverAtDoorIso": driver_at_door,
        "unloadStartIso": unload_start,
        "unloadFinishIso": unload_finish,
        "receivingStartIso": receiving_start,
        "receivingFinishIso": receiving_finish,
        "driverLeftIso": driver_left,
        "finishIso": unload_finish or receiving_finish,
        "statusText": str(merged.get("drstat") or "").strip(),
    }


def _merge_dms_portal_rows(loads: List[Dict[str, Any]], stamps: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    stamps_by_key: Dict[str, Dict[str, Any]] = {}
    for stamp in stamps:
        key = _dms_key(stamp)
        if key:
            stamps_by_key[key] = stamp
    seen = set()
    trucks = []

    def _keep(truck: Dict[str, Any]) -> bool:
        # Keep every real, non-rejected truck for the shift (whether or not it has
        # checked in yet) so shift progress can be measured against the FULL set.
        # The endpoint hides not-yet-checked-in trucks from the board separately.
        if str(truck.get("statusText") or "").strip().lower() == "rejected":
            return False  # rejected loads — we won't be unloading those trucks
        return bool(truck["checkInIso"] or truck["door"] or truck["ref"] or truck["po"])

    for load in loads:
        key = _dms_key(load)
        if key:
            seen.add(key)
        truck = _normalize_portal_truck(load, stamps_by_key.get(key, {}))
        if _keep(truck):
            trucks.append(truck)
    for stamp in stamps:
        key = _dms_key(stamp)
        if key and key in seen:
            continue
        truck = _normalize_portal_truck({}, stamp)
        if _keep(truck):
            trucks.append(truck)
    return trucks


def _learn_from_dms(trucks: List[Dict[str, Any]], shift_date: str) -> int:
    """Record each completed truck's real turn time (check-in -> unload finish)
    per vendor so the completion estimate gets more accurate over time.
    Idempotent per (vendor, truck_ref, shift_date) so repeated pulls of the same
    shift don't double-count. Best-effort; never raises into the request."""
    inserted = 0
    try:
        conn = get_db()
        now_str = datetime.now(timezone.utc).isoformat()
        for t in trucks:
            vendor = (t.get("supplier") or "").strip().upper()
            ci = t.get("checkInIso")
            done = t.get("unloadFinishIso") or t.get("receivingFinishIso")
            ref = str(t.get("rowid") or t.get("ref") or "").strip()
            if not vendor or not ci or not done or not ref:
                continue
            try:
                dock_min = round((datetime.fromisoformat(done) - datetime.fromisoformat(ci)).total_seconds() / 60)
            except Exception:
                continue
            if dock_min <= 0 or dock_min > 720:
                continue
            exists = conn.execute(
                "SELECT 1 FROM vendor_unload_times WHERE vendor=? AND truck_ref=? AND shift_date=? LIMIT 1",
                (vendor, ref, shift_date),
            ).fetchone()
            if exists:
                continue
            conn.execute(
                "INSERT INTO vendor_unload_times (vendor, dock_min, shift_date, source, truck_ref, recorded_at) "
                "VALUES (?,?,?,?,?,?)",
                (vendor, dock_min, shift_date, "DMS", ref, now_str),
            )
            inserted += 1
        conn.commit()
        conn.close()
    except Exception:
        pass
    return inserted


# ── Auth helpers ──────────────────────────────────────────────────────────────

def require_auth(session: Optional[str] = Cookie(default=None)) -> str:
    if not session or session not in _sessions:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return _sessions[session]


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.post("/api/login")
def login(creds: LoginIn, response: Response):
    pw_hash = _hash(creds.password)
    stored  = _USERS.get(creds.username.strip().lower())
    if stored is None or stored != pw_hash:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    token = secrets.token_hex(32)
    uname = creds.username.strip().lower()
    _sessions[token] = uname
    response.set_cookie(
        "session", token,
        httponly=True,
        samesite="lax",
        max_age=60 * 60 * 24 * 7,   # 7 days
        secure=os.environ.get("RAILWAY_ENVIRONMENT") is not None,
    )
    # Record visit for daily counter
    try:
        conn = get_db()
        now = datetime.now(timezone.utc).isoformat()
        conn.execute("INSERT INTO visits (username, visited_at) VALUES (?, ?)", (uname, now))
        conn.commit()
        conn.close()
    except Exception:
        pass
    return {"ok": True, "username": uname, "role": _ROLES.get(uname, "guest")}


@app.post("/api/logout")
def logout(response: Response, session: Optional[str] = Cookie(default=None)):
    if session and session in _sessions:
        del _sessions[session]
    response.delete_cookie("session")
    return {"ok": True}


@app.get("/api/me")
def me(username: str = Depends(require_auth)):
    return {"username": username, "role": _ROLES.get(username, "guest")}


@app.get("/api/daily-visitors")
def daily_visitors(username: str = Depends(require_auth)):
    """Return count of logins in the last 24 hours."""
    try:
        conn = get_db()
        cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        row = conn.execute(
            "SELECT COUNT(*) as cnt FROM visits WHERE visited_at >= ?", (cutoff,)
        ).fetchone()
        conn.close()
        return {"count": row["cnt"] if row else 0}
    except Exception:
        return {"count": 0}


# ── Reports ───────────────────────────────────────────────────────────────────

@app.post("/api/reports", status_code=201)
def create_report(report: ReportIn, _: str = Depends(require_auth)):
    conn = get_db()
    try:
        cur = conn.execute(
            """INSERT INTO reports (filename, uploaded_at, row_count, headers, rows, selectors)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                report.filename,
                datetime.now(timezone.utc).isoformat(),
                len(report.rows),
                json.dumps(report.headers),
                json.dumps(report.rows),
                json.dumps(report.selectors),
            ),
        )
        rid = cur.lastrowid
        conn.commit()
    finally:
        conn.close()
    return {"id": rid, "filename": report.filename, "row_count": len(report.rows)}


@app.get("/api/reports")
def list_reports(_: str = Depends(require_auth)):
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT id, filename, uploaded_at, row_count FROM reports ORDER BY uploaded_at DESC"
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


@app.get("/api/reports/{report_id}")
def get_report(report_id: int, _: str = Depends(require_auth)):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Report not found")
    return {
        "id":          row["id"],
        "filename":    row["filename"],
        "uploaded_at": row["uploaded_at"],
        "row_count":   row["row_count"],
        "headers":     json.loads(row["headers"]),
        "rows":        json.loads(row["rows"]),
        "selectors":   json.loads(row["selectors"]),
    }


@app.delete("/api/reports/{report_id}")
def delete_report(report_id: int, _: str = Depends(require_auth)):
    conn = get_db()
    try:
        conn.execute("DELETE FROM reports WHERE id = ?", (report_id,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


# ── Audit decisions ───────────────────────────────────────────────────────────

@app.get("/api/decisions")
def list_decisions(report_id: Optional[int] = None, _: str = Depends(require_auth)):
    conn = get_db()
    try:
        if report_id is not None:
            rows = conn.execute(
                "SELECT * FROM audit_decisions WHERE report_id = ? ORDER BY decided_at DESC",
                (report_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM audit_decisions ORDER BY decided_at DESC"
            ).fetchall()
    finally:
        conn.close()
    return [
        {**dict(r), "po_keys": json.loads(r["po_keys"] or "[]")}
        for r in rows
    ]


@app.post("/api/decisions")
def save_decision(d: DecisionIn, _: str = Depends(require_auth)):
    conn = get_db()
    try:
        conn.execute(
            """
            INSERT INTO audit_decisions
                (truck_key, decision, report_id, po_keys, supplier_key, decided_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(truck_key) DO UPDATE SET
                decision     = excluded.decision,
                report_id    = excluded.report_id,
                po_keys      = excluded.po_keys,
                supplier_key = excluded.supplier_key,
                decided_at   = excluded.decided_at
            """,
            (
                d.truck_key,
                d.decision,
                d.report_id,
                json.dumps(d.po_keys or []),
                d.supplier_key,
                datetime.now(timezone.utc).isoformat(),
            ),
        )
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.delete("/api/decisions")
def delete_decision(truck_key: str, _: str = Depends(require_auth)):
    conn = get_db()
    try:
        conn.execute("DELETE FROM audit_decisions WHERE truck_key = ?", (truck_key,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}




# ── Frontend ──────────────────────────────────────────────────────────────────


def _build_dms_portal_payload(session: Dict[str, Any], date: Optional[str]) -> Dict[str, Any]:
    info = _dms_business_date(date)
    base_payload = {
        "info": info,
        "loc": session["loc"],
        "userinfo": session["userinfo"],
        "buck": session.get("buck") or {},
    }
    loads_response = _dms_json_request("api/load/getloaddetails", base_payload, session["config"])
    stamps_response = _dms_json_request("api/stamp/getStamps", base_payload, session["config"])
    loads = [x for x in _first_list(loads_response) if isinstance(x, dict)]
    stamps = [x for x in _first_list(stamps_response) if isinstance(x, dict)]
    all_trucks = _merge_dms_portal_rows(loads, stamps)
    board = [t for t in all_trucks if t.get("checkInIso")]

    def _stat(st):
        return str(st.get("drstat") or "").strip().lower()

    rejected_count = sum(1 for st in stamps if _stat(st) == "rejected")
    no_show_count = sum(
        1 for st in stamps
        if _stat(st) == "late" and not (st.get("drchk") or st.get("clrkchk"))
    )

    now_utc = datetime.now(timezone.utc)
    no_show_trucks = []
    for t in all_trucks:
        if t.get("checkInIso"):
            continue
        appt_iso = t.get("appointmentIso")
        if not appt_iso:
            continue
        try:
            appt_dt = datetime.fromisoformat(appt_iso)
            if appt_dt.tzinfo is None:
                appt_dt = appt_dt.replace(tzinfo=timezone.utc)
            if appt_dt < now_utc:
                no_show_trucks.append({
                    "ref": t.get("ref", ""),
                    "supplier": t.get("supplier", ""),
                    "door": t.get("door", ""),
                    "appointmentIso": appt_iso,
                })
        except Exception:
            pass

    area_summary: Dict[str, Dict[str, Any]] = {}
    for t in all_trucks:
        a = (t.get("area") or "").strip() or "-"
        s = area_summary.setdefault(a, {"area": a, "scheduled": 0, "checked_in": 0, "completed": 0})
        s["scheduled"] += 1
        if t.get("checkInIso"):
            s["checked_in"] += 1
        if t.get("unloadFinishIso") or t.get("receivingFinishIso"):
            s["completed"] += 1
    area_list = sorted(area_summary.values(), key=lambda x: -x["scheduled"])

    door_set = set()
    for t in all_trucks:
        dd = str(t.get("door") or "").strip()
        if dd.isdigit():
            door_set.add(int(dd))

    return {
        "ok": True,
        "business_date": info,
        "location": session["loc"],
        "load_count": len(loads),
        "stamp_count": len(stamps),
        "trucks": board,
        "scheduled_trucks": all_trucks,
        "total_expected": len(all_trucks),
        "checked_in_count": len(board),
        "rejected_count": rejected_count,
        "no_show_count": no_show_count,
        "no_show_trucks": no_show_trucks,
        "area_summary": area_list,
        "doors": sorted(door_set),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


@app.get("/api/dms/session")
def dms_session_status():
    session = _ensure_dms_session()
    loc = session.get("loc") or {}
    return {
        "ok": True,
        "location": {
            "locid": loc.get("locid"),
            "name": loc.get("name") or loc.get("locName") or loc.get("location"),
            "cCode": loc.get("cCode") or loc.get("code"),
        },
        "cached_at": session.get("cached_at"),
    }


@app.get("/api/dms/mn/session")
def dms_mn_session_status(force: bool = False):
    session = _ensure_dms_mn_session(force=force)
    loc = session.get("loc") or {}
    return {
        "ok": True,
        "location": {
            "locid": loc.get("locid"),
            "name": loc.get("name") or loc.get("locName") or loc.get("location"),
            "cCode": loc.get("cCode") or loc.get("code"),
        },
        "cached_at": session.get("cached_at"),
    }


@app.get("/api/dms/mn/portal")
def dms_mn_portal(date: Optional[str] = None, force: bool = False):
    """Read Minnesota DMS load/stamp rows for the MN My Portal. This route never writes to DMS."""
    session = _ensure_dms_mn_session(force=force)
    return _build_dms_portal_payload(session, date)


@app.get("/api/dms/portal")
def dms_portal(date: Optional[str] = None, force: bool = False, debug: bool = False):
    """Read DMS load/stamp rows for My Portal. This route never writes to DMS."""
    # In debug mode, never 500 — capture and return whatever we can learn.
    if debug:
        dbg: Dict[str, Any] = {}
        try:
            session = _ensure_dms_session(force=force)
        except Exception as e:
            return {"ok": False, "debug": {"stage": "login", "error": str(e)}}
        info = _dms_business_date(date)
        base_payload = {
            "info": info, "loc": session["loc"],
            "userinfo": session["userinfo"], "buck": session.get("buck") or {},
        }
        dbg["business_date"] = info
        dbg["location"] = session["loc"]
        try:
            loads_response = _dms_json_request("api/load/getloaddetails", base_payload, session["config"])
            loads = [x for x in _first_list(loads_response) if isinstance(x, dict)]
        except Exception as e:
            loads, dbg["loads_error"] = [], str(e)
        try:
            stamps_response = _dms_json_request("api/stamp/getStamps", base_payload, session["config"])
            stamps = [x for x in _first_list(stamps_response) if isinstance(x, dict)]
        except Exception as e:
            stamps, dbg["stamps_error"] = [], str(e)
        trucks = _merge_dms_portal_rows(loads, stamps)
        dbg.update({
            "load_count": len(loads),
            "stamp_count": len(stamps),
            "load_keys": sorted(loads[0].keys()) if loads else [],
            "stamp_keys": sorted(stamps[0].keys()) if stamps else [],
            "sample_load": loads[0] if loads else None,
            "sample_stamp": stamps[0] if stamps else None,
            "first_truck_normalized": trucks[0] if trucks else None,
            "truck_count": len(trucks),
        })
        return {"ok": True, "debug": dbg}

    session = _ensure_dms_session(force=force)
    info = _dms_business_date(date)
    base_payload = {
        "info": info,
        "loc": session["loc"],
        "userinfo": session["userinfo"],
        "buck": session.get("buck") or {},
    }
    loads_response = _dms_json_request("api/load/getloaddetails", base_payload, session["config"])
    stamps_response = _dms_json_request("api/stamp/getStamps", base_payload, session["config"])
    loads = [x for x in _first_list(loads_response) if isinstance(x, dict)]
    stamps = [x for x in _first_list(stamps_response) if isinstance(x, dict)]
    all_trucks = _merge_dms_portal_rows(loads, stamps)
    # Board shows only trucks physically here (checked in). Not-yet-checked-in
    # trucks stay off the board but still count toward total_expected, so shift
    # progress is measured against the whole shift, not just what's on the dock.
    board = [t for t in all_trucks if t.get("checkInIso")]
    # Learn from completed trucks automatically (deduped) so the completion
    # estimate sharpens over time with zero manual save.
    _learn_from_dms(board, info)
    # Counts for the end-of-shift report.
    def _stat(st):
        return str(st.get("drstat") or "").strip().lower()
    rejected_count = sum(1 for st in stamps if _stat(st) == "rejected")
    no_show_count = sum(
        1 for st in stamps
        if _stat(st) == "late" and not (st.get("drchk") or st.get("clrkchk"))
    )
    # Trucks with a past appointment but still not checked in — shown in the
    # no-show tracker on the frontend.
    now_utc = datetime.now(timezone.utc)
    no_show_trucks = []
    for t in all_trucks:
        if t.get("checkInIso"):
            continue
        appt_iso = t.get("appointmentIso")
        if not appt_iso:
            continue
        try:
            appt_dt = datetime.fromisoformat(appt_iso)
            if appt_dt.tzinfo is None:
                appt_dt = appt_dt.replace(tzinfo=timezone.utc)
            if appt_dt < now_utc:
                no_show_trucks.append({
                    "ref":            t.get("ref", ""),
                    "supplier":       t.get("supplier", ""),
                    "door":           t.get("door", ""),
                    "appointmentIso": appt_iso,
                })
        except Exception:
            pass
    # Per-area summary across ALL scheduled (non-rejected) trucks — feeds both the
    # shift report and the Area Breakdown view.
    area_summary: Dict[str, Dict[str, Any]] = {}
    for t in all_trucks:
        a = (t.get("area") or "").strip() or "—"
        s = area_summary.setdefault(a, {"area": a, "scheduled": 0, "checked_in": 0, "completed": 0})
        s["scheduled"] += 1
        if t.get("checkInIso"):
            s["checked_in"] += 1
        if t.get("unloadFinishIso") or t.get("receivingFinishIso"):
            s["completed"] += 1
    area_list = sorted(area_summary.values(), key=lambda x: -x["scheduled"])
    # Every door used anywhere in today's schedule (so the door map can show the
    # full dock, not just doors with a truck on them right now).
    door_set = set()
    for t in all_trucks:
        dd = str(t.get("door") or "").strip()
        if dd.isdigit():
            door_set.add(int(dd))
    doors = sorted(door_set)
    return {
        "ok": True,
        "business_date": info,
        "location": session["loc"],
        "load_count": len(loads),
        "stamp_count": len(stamps),
        "trucks": board,
        "total_expected": len(all_trucks),
        "checked_in_count": len(board),
        "rejected_count": rejected_count,
        "no_show_count": no_show_count,
        "no_show_trucks": no_show_trucks,
        "area_summary": area_list,
        "doors": doors,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


@app.post("/api/portal/learn-history")
def portal_learn_history(days: int = 7):
    """Backfill vendor learning from the last N days of real DMS shifts so the
    completion estimate is accurate right away instead of only over time."""
    session = _ensure_dms_session()
    days = max(1, min(int(days or 7), 30))
    total_learned = 0
    dates_done = []
    base_loc = session["loc"]
    base_user = session["userinfo"]
    base_buck = session.get("buck") or {}
    today = datetime.now()
    for back in range(days):
        d = today - timedelta(days=back)
        info = f"{d.month}/{d.day}/{d.year}"
        payload = {"info": info, "loc": base_loc, "userinfo": base_user, "buck": base_buck}
        try:
            loads = [x for x in _first_list(_dms_json_request("api/load/getloaddetails", payload, session["config"])) if isinstance(x, dict)]
            stamps = [x for x in _first_list(_dms_json_request("api/stamp/getStamps", payload, session["config"])) if isinstance(x, dict)]
            trucks = _merge_dms_portal_rows(loads, stamps)
            learned = _learn_from_dms(trucks, info)
            total_learned += learned
            dates_done.append({"date": info, "trucks": len(trucks), "learned": learned})
        except Exception as e:
            dates_done.append({"date": info, "error": str(e)})
    return {"ok": True, "days": days, "total_learned": total_learned, "by_date": dates_done}

class DmsStampIn(BaseModel):
    load_id: Optional[str] = None
    po: Optional[str] = None
    stamp_type: str
    stamp_time: Optional[str] = None


class DmsScheduleRowIn(BaseModel):
    row_number: Optional[int] = None
    dms_truck_number: Optional[int] = None
    reference: str = ""
    po: str = ""
    scheduled_text: str = ""
    scheduled_iso: str = ""
    pallets: Optional[int] = None
    quantity: str = ""
    supplier: str = ""
    dock: str = ""
    protection: str = ""
    product_category: str = ""
    incoterm: str = ""
    current_state: str = ""
    area: str = ""
    dock_type: str = ""


class DmsScheduleUploadIn(BaseModel):
    business_date: Optional[str] = None
    rows: List[DmsScheduleRowIn]
    skip_existing: bool = True
    dry_run: bool = False

STAMP_TYPE_MAP = {
    "checkin":          "checkIn",
    "check_in":         "checkIn",
    "driveratdoor":     "driverAtDoor",
    "driver_at_door":   "driverAtDoor",
    "unloadstart":      "unloadStart",
    "unload_start":     "unloadStart",
    "unloadfinish":     "unloadFinish",
    "unload_finish":    "unloadFinish",
    "receivingfinish":  "receivingFinish",
    "receiving_finish": "receivingFinish",
}


def _schedule_digits(value: Any) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _schedule_po_keys(value: Any) -> List[str]:
    text = str(value or "")
    keys = []
    for part in text.replace(";", ",").split(","):
        digits = _schedule_digits(part)
        if digits:
            keys.append(digits)
    if not keys:
        digits = _schedule_digits(text)
        if digits:
            keys.append(digits)
    return list(dict.fromkeys(keys))


def _schedule_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return max(0, int(float(str(value).strip())))
    except Exception:
        return default


def _schedule_appt_to_utc_string(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        raise ValueError("scheduled time is missing")
    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except Exception:
        parsed = None
        for fmt in ("%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                parsed = datetime.strptime(text, fmt)
                break
            except ValueError:
                pass
        if parsed is None:
            raise ValueError(f"scheduled time is unreadable: {text}")
        dt = parsed.replace(tzinfo=DMS_BUSINESS_TZ)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=DMS_BUSINESS_TZ)
    return dt.astimezone(timezone.utc).strftime("%a, %d %b %Y %H:%M:%S GMT")


def _area_matches_hint(area: Any, hint: str) -> bool:
    area_text = " ".join(str(v) for v in area.values()) if isinstance(area, dict) else str(area or "")
    area_text = area_text.lower()
    hint = hint.lower()
    target = hint.split("|", 1)[0].strip()
    if target == "cold plants" and any(word in area_text for word in ("cold", "cooler", "chill", "plant", "produce")):
        return True
    if target == "plants" and any(word in area_text for word in ("plant", "produce", "floral")):
        return True
    if target == "produce" and any(word in area_text for word in ("produce", "fruit", "vegetable", "veggie", "veg")):
        return True
    if target and target not in ("cold plants", "plants"):
        return target in area_text
    pairs = [
        ("freezer", ("freezer", "frozen", "freeze")),
        ("produce", ("produce", "fruit", "fruits", "veg", "vegetable", "vegetables", "veggie", "veggies", "salad")),
        ("cooler", ("cooler", "chill", "chiller", "meat", "dairy", "refrigerated")),
        ("dry", ("dry", "ambient", "grocery", "pantry", "beverage", "household")),
    ]
    for area_key, words in pairs:
        if any(w in hint for w in words) and area_key in area_text:
            return True
    return False


def _schedule_match_text(value: Any) -> str:
    return " ".join("".join(ch.lower() if ch.isalnum() else " " for ch in str(value or "")).split())


_SCHEDULE_TALL_DOOR_THRESHOLD = 0.68
_SCHEDULE_MANUAL_DOOR_RULES = (
    (("sonstegard", "sunrisefarm", "sunrise farm"), True),
    (("frito lay",), False),
    (("niagara", "niagara bottling"), False),
    (("cactus",), True),
    (("taylor farms retail",), True),
    (("taylor farms texas",), False),
    (("bonipak", "boni pak"), True),
    (("simply fresh",), True),
    (("ajm packaging", "snack king", "snak king", "post consumer brands", "aspen", "schulze", "birch"), True),
    (("mowi",), False),
    (("great lakes cheese",), True),
    (("cafe valley",), True),
    (("columbia fruit", "richelieu foods", "la fournee", "pdm vegetables", "bimbo bakehouse"), True),
    (("del monte",), False),
    (("ns brands", "ns brand"), True),
    (("absopure", "absopure water"), False),
    (("saputo cheese", "saputo"), False),
    (("southern corporate packing", "southern corp packing", "southern corp packers", "southern corp packers inc"), False),
    (("ganfer", "ganfer fresh"), False),
    (("aurora organic", "aurora organic dairy"), False),
    (("schrieber", "schreiber"), False),
    (("tanimura and antle", "tanimura antle fresh foods"), True),
)
_SCHEDULE_TALL_HISTORY_CACHE: Optional[List[Dict[str, Any]]] = None


def _schedule_supplier_key(value: Any) -> str:
    text = str(value or "").lower().replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\b(inc|incorporated|llc|ltd|limited|co|corp|corporation|company|lp|plc|usa|us)\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _schedule_manual_tall_status(supplier: Any) -> Optional[bool]:
    key = _schedule_supplier_key(supplier)
    if not key:
        return None
    if key == "ns" or key.startswith("ns ") or "ns brand" in key:
        return True
    for patterns, needs_tall in _SCHEDULE_MANUAL_DOOR_RULES:
        for pattern in patterns:
            rule_key = _schedule_supplier_key(pattern)
            if rule_key and (rule_key in key or key in rule_key):
                return needs_tall
    return None


def _schedule_tall_history_rows() -> List[Dict[str, Any]]:
    global _SCHEDULE_TALL_HISTORY_CACHE
    if _SCHEDULE_TALL_HISTORY_CACHE is not None:
        return _SCHEDULE_TALL_HISTORY_CACHE
    rows: List[Dict[str, Any]] = []
    try:
        text = HISTORICAL_DATA_PATH.read_text(encoding="utf-8")
        match = re.search(r'"tallDoorSuppliers"\s*:\s*(\[.*?\])\s*,\s*"doorRanges"', text, re.S)
        if match:
            parsed = json.loads(match.group(1))
            if isinstance(parsed, list):
                for item in parsed:
                    if isinstance(item, dict):
                        key = _schedule_supplier_key(item.get("key"))
                        if key:
                            rows.append({**item, "normalized_key": key})
    except Exception:
        rows = []
    _SCHEDULE_TALL_HISTORY_CACHE = rows
    return rows


def _schedule_tall_history_for(supplier: Any) -> Optional[Dict[str, Any]]:
    key = _schedule_supplier_key(supplier)
    if not key:
        return None
    compact_key = key.replace(" ", "")
    for item in _schedule_tall_history_rows():
        history_key = str(item.get("normalized_key") or "")
        compact_history = history_key.replace(" ", "")
        if key == history_key:
            return item
        if len(key) >= 8 and (history_key in key or key in history_key):
            return item
        if len(compact_key) >= 8 and (compact_history in compact_key or compact_key in compact_history):
            return item
    return None


_SCHEDULE_MULTI_PO_TALL_PALLET_THRESHOLD = 26


def _schedule_needs_tall_door(row: DmsScheduleRowIn) -> bool:
    manual = _schedule_manual_tall_status(row.supplier)
    if manual is not None:
        return manual
    # Multi-PO loads with no supplier listed and a heavy pallet count need a
    # tall door (confirmed with James, OKS-only).
    if not row.supplier.strip() and len(_schedule_po_keys(row.po)) > 1:
        if _schedule_int(row.pallets, 0) > _SCHEDULE_MULTI_PO_TALL_PALLET_THRESHOLD:
            return True
    history = _schedule_tall_history_for(row.supplier)
    if not history:
        return False
    total = _schedule_int(history.get("total"), 0)
    tall = _schedule_int(history.get("tall"), 0)
    return bool(total and tall / total >= _SCHEDULE_TALL_DOOR_THRESHOLD)


_SCHEDULE_BREAKDOWN_HISTORY_CACHE: Optional[List[Dict[str, Any]]] = None


def _schedule_breakdown_history_rows() -> List[Dict[str, Any]]:
    global _SCHEDULE_BREAKDOWN_HISTORY_CACHE
    if _SCHEDULE_BREAKDOWN_HISTORY_CACHE is not None:
        return _SCHEDULE_BREAKDOWN_HISTORY_CACHE
    rows: List[Dict[str, Any]] = []
    try:
        text = HISTORICAL_DATA_PATH.read_text(encoding="utf-8")
        match = re.search(r'"breakdownSuppliers"\s*:\s*(\[.*?\])\s*,\s*"tallDoorSuppliers"', text, re.S)
        if match:
            parsed = json.loads(match.group(1))
            if isinstance(parsed, list):
                for item in parsed:
                    if isinstance(item, dict):
                        key = _schedule_supplier_key(item.get("key"))
                        if key:
                            rows.append({**item, "normalized_key": key})
    except Exception:
        rows = []
    _SCHEDULE_BREAKDOWN_HISTORY_CACHE = rows
    return rows


def _schedule_breakdown_history_for(supplier: Any) -> Optional[Dict[str, Any]]:
    key = _schedule_supplier_key(supplier)
    if not key:
        return None
    compact_key = key.replace(" ", "")
    for item in _schedule_breakdown_history_rows():
        history_key = str(item.get("normalized_key") or "")
        compact_history = history_key.replace(" ", "")
        if key == history_key:
            return item
        if len(key) >= 8 and (history_key in key or key in history_key):
            return item
        if len(compact_key) >= 8 and (compact_history in compact_key or compact_key in compact_history):
            return item
    return None


def _schedule_needs_breakdown(row: DmsScheduleRowIn) -> bool:
    history = _schedule_breakdown_history_for(row.supplier)
    if not history:
        return False
    total = _schedule_int(history.get("total"), 0)
    breakdowns = _schedule_int(history.get("breakdowns"), 0)
    return bool(total and breakdowns / total >= _SCHEDULE_TALL_DOOR_THRESHOLD)


def _schedule_tall_door_note(row: DmsScheduleRowIn) -> str:
    tall = "YES" if _schedule_needs_tall_door(row) else "NO"
    breakdown = "YES" if _schedule_needs_breakdown(row) else "NO"
    return f"Tall Door: {tall} / Breakdown: {breakdown}"


def _schedule_dock_type(row: DmsScheduleRowIn, use_oks_rules: bool = True) -> str:
    supplier_text = (row.supplier or "").lower()
    supplier_key = _schedule_match_text(row.supplier)
    dock_category_text = " ".join([
        row.dock or "",
        row.product_category or "",
    ]).lower()
    all_text = " ".join([
        row.supplier or "",
        row.area or "",
        row.protection or "",
        row.dock or "",
        row.product_category or "",
    ]).lower()
    dry_supplier_names = (
        "clasen quality chocolate",
        "perfetti van melle usa",
        "interbake foods",
        "silvestri sweets",
        "the hershey company",
        "hershey",
        "gilster mary lee",
    )
    if use_oks_rules and any(name in supplier_key for name in dry_supplier_names):
        return "Dry"
    if use_oks_rules and "falcon" in supplier_key and "farm" in supplier_key:
        return "Cold Plants"
    if use_oks_rules and (any(word in dock_category_text for word in ("plant", "plants", "floral", "flower")) or any(word in supplier_text for word in ("plant", "plants", "floral", "flower"))):
        if any(word in all_text for word in ("cold", "cooler", "chill", "refrigerated")) and "falcon" in supplier_text:
            return "Cold Plants"
        return "Plants"
    if any(word in dock_category_text for word in ("produce", "fruit", "fruits", "veg", "vegetable", "vegetables", "veggie", "veggies", "salad")):
        return "Produce"
    if any(word in all_text for word in ("freezer", "frozen", "freeze")):
        return "Freezer"
    if any(word in all_text for word in ("cooler", "chill", "chiller", "dairy", "meat", "refrigerated")):
        return "Cooler"
    if any(word in all_text for word in ("dry", "ambient", "grocery", "pantry", "beverage", "household")):
        return "Dry"
    return row.dock_type or row.area or row.dock or ""


# Minnesota schedule uploader — deterministic area mapping ported from Orbit's
# PowerView ALDI schedule parser (orbit/app.py::_parse_aldi_schedule /
# _DMS_AREA_MAP). Confirmed 2026-09 against Minnesota's live DMS area list:
# AMB, CHL, EGGS, FLOOR LOADED, FRESH MEAT, FRZ, INTERNATIONAL,
# PLANTS/FLOWERS, PRODUCE, SLIP SHEET.
_MN_DMS_AREA_MAP: Dict[Any, str] = {
    ("ambient", "std. trailer dry"):     "AMB",
    ("ambient", "floor loaded"):         "FLOOR LOADED",
    ("ambient", "slip sheet"):           "SLIP SHEET",
    ("chiller", "std. trailer cooler"):  "CHL",
    ("chiller", "produce"):              "PRODUCE",
    ("chiller", "fresh meat"):           "FRESH MEAT",
    ("chiller", "eggs"):                 "EGGS",
    ("freezer", "std. trailer freezer"): "FRZ",
}
_MN_DMS_PROT_FALLBACK: Dict[str, str] = {"ambient": "AMB", "chiller": "CHL", "freezer": "FRZ"}
# Dock category alone can also signal these two areas, regardless of protection
# level (confirmed with James: dock type "International" / "Plants/Flowers").
_MN_DMS_DOCK_ONLY_MAP: Dict[str, str] = {
    "international":   "INTERNATIONAL",
    "plants/flowers":   "PLANTS/FLOWERS",
    "plants":           "PLANTS/FLOWERS",
    "flowers":          "PLANTS/FLOWERS",
}


def _mn_schedule_area(row: DmsScheduleRowIn) -> Optional[str]:
    dock_key = (row.dock or row.product_category or "").strip().lower()
    if dock_key in _MN_DMS_DOCK_ONLY_MAP:
        return _MN_DMS_DOCK_ONLY_MAP[dock_key]
    prot_key = (row.protection or "").strip().lower()
    mapped = _MN_DMS_AREA_MAP.get((prot_key, dock_key))
    if mapped:
        return mapped
    return _MN_DMS_PROT_FALLBACK.get(prot_key)


def _dms_area_for_schedule(row: DmsScheduleRowIn, session: Dict[str, Any], business_date: str, use_oks_rules: bool = True) -> Any:
    if not use_oks_rules:
        # Minnesota: try the deterministic PowerView-style table first; only
        # fall through to the fuzzy live-areas hint match below for a
        # protection/dock combination the table doesn't cover.
        mn_area = _mn_schedule_area(row)
        if mn_area:
            return mn_area
    buck = session.get("buck") or {}
    areas = buck.get("areas") if isinstance(buck, dict) else None
    if not areas:
        try:
            buck = _dms_json_request(
                "api/dash/getbucketdata",
                {"info": business_date, "loc": session["loc"]},
                session["config"],
            )
            if isinstance(buck, dict):
                session["buck"] = buck
                areas = buck.get("areas")
        except Exception:
            areas = None
    if isinstance(areas, list) and areas:
        dock_type = _schedule_dock_type(row, use_oks_rules)
        hint = " | " + " ".join([row.area, row.protection, row.dock, row.product_category, row.supplier])
        for area in areas:
            if _area_matches_hint(area, dock_type.lower() + hint):
                return area
        if dock_type in ("Produce", "Plants", "Cold Plants"):
            return dock_type
        return areas[0]
    return _schedule_dock_type(row, use_oks_rules)


def _dms_schedule_upload_model(session: Dict[str, Any]) -> Dict[str, Any]:
    candidates = [
        {"user": {"user": session["userinfo"]}},
        {"user": session["userinfo"]},
    ]
    last_error = None
    for payload in candidates:
        try:
            result = _dms_json_request("api/load/getuploadmodel", payload, session["config"])
            if isinstance(result, dict):
                return result
        except Exception as exc:
            last_error = str(exc)
    return {"_error": last_error or "DMS upload model was not returned."}


def _powerview_strip_commodity_code(value: str) -> str:
    """PowerView's cleanup (orbit/app.py::_strip_commodity_code): the ALDI
    schedule codes commodities as "0008 Fruits & Vegetables"; DMS wants just
    "Fruits & Vegetables". Only strips a leading run of digits + whitespace."""
    text = str(value or "").strip()
    return re.sub(r"^\s*\d+\s+", "", text) if text else ""


def _powerview_incoterm_tag_supplier(supplier: str, incoterm: str) -> str:
    """PowerView's incoterm 'vendor' tag mode (orbit/app.py::_apply_incoterm_tag):
    prefix the supplier with its Incoterm code in parentheses, e.g.
    "(DDP) Waterloo Sparkling Water Corp." or "(EXW) Some Vendor Inc."."""
    code = str(incoterm or "").strip().upper()
    if not code:
        return supplier
    return f"({code}) {supplier}".strip()


def _dms_schedule_insert_payload(
    row: DmsScheduleRowIn,
    session: Dict[str, Any],
    business_date: str,
    truck_number: int,
    upload_model: Optional[Dict[str, Any]] = None,
    use_oks_rules: bool = True,
) -> Dict[str, Any]:
    notes = _schedule_tall_door_note(row) if use_oks_rules else "Imported from schedule"

    base_insert: Dict[str, Any] = {}
    if isinstance(upload_model, dict) and isinstance(upload_model.get("insert"), dict):
        base_insert.update(upload_model["insert"])

    dock_type = _schedule_dock_type(row, use_oks_rules)
    # Keep the full PO string intact (even a multi-PO cell like
    # "7521444135,7522401186") -- that's what lets My Portal's search box find
    # the truck by typing any one of its PO numbers, same as OKS today.
    po_number = row.po.strip()
    category_desc = row.product_category.strip() if use_oks_rules else _powerview_strip_commodity_code(row.product_category)
    supplier_fallback = "MULTI PO SUPPLIER NOT KNOWN" if use_oks_rules else "MULTI PO - VENDOR NOT SPECIFIED"
    supplier_value = row.supplier.strip() or supplier_fallback
    supplier_value = _powerview_incoterm_tag_supplier(supplier_value, row.incoterm)
    base_insert.update({
        "area": _dms_area_for_schedule(row, session, business_date, use_oks_rules),
        "poNum": po_number,
        "appt": _schedule_appt_to_utc_string(row.scheduled_iso or row.scheduled_text),
        "trkNum": max(1, int(truck_number or 1)),
        "doorNum": 0,
        "load": dock_type or row.protection.strip() or row.dock.strip() or "ALDI Schedule",
        "sup": supplier_value,
        "qty": _schedule_int(row.pallets, 0),
        "carr": str(base_insert.get("carr") or ""),
        "notes": notes,
        "desc": category_desc or dock_type or row.dock.strip(),
        "cabNum": str(base_insert.get("cabNum") or ""),
    })

    return {
        "ins": base_insert,
        "loc": session["loc"],
        "tzOffset": int(datetime.now().astimezone().utcoffset().total_seconds() / -60),
        "busDate": business_date,
    }


def _dms_success(result: Any) -> bool:
    if isinstance(result, dict):
        if result.get("operationSuccess") is False or result.get("ok") is False or result.get("success") is False:
            return False
        if result.get("error"):
            return False
    return True


def _run_dms_schedule_upload(session: Dict[str, Any], body: DmsScheduleUploadIn, use_oks_rules: bool = True) -> Dict[str, Any]:
    if not body.rows:
        raise HTTPException(status_code=400, detail="No schedule rows were provided.")
    if len(body.rows) > 300:
        raise HTTPException(status_code=400, detail="Schedule upload is limited to 300 rows at a time.")

    business_date = _dms_business_date(body.business_date)
    upload_model = _dms_schedule_upload_model(session)
    model_error = upload_model.get("_error") if isinstance(upload_model, dict) else "DMS upload model was not returned."
    existing_keys = set()
    if body.skip_existing:
        try:
            existing_payload = {
                "info": business_date,
                "loc": session["loc"],
                "userinfo": session["userinfo"],
                "buck": session.get("buck") or {},
            }
            existing_loads = [
                x for x in _first_list(_dms_json_request("api/load/getloaddetails", existing_payload, session["config"]))
                if isinstance(x, dict)
            ]
            for item in existing_loads:
                for key in _schedule_po_keys(item.get("poNum") or item.get("po")):
                    existing_keys.add(f"po:{key}")
        except Exception:
            existing_keys = set()

    results = []
    inserted = skipped = failed = 0
    next_truck_number = 1
    truck_number_by_reference: Dict[str, int] = {}
    for index, row in enumerate(body.rows, start=1):
        po_keys = _schedule_po_keys(row.po)
        row_id = row.row_number or index
        reference_key = _schedule_digits(row.reference) or row.reference.strip().upper()
        truck_number = 0
        try:
            if not po_keys:
                raise ValueError("PO number is missing")
            if not row.scheduled_iso and not row.scheduled_text:
                raise ValueError("scheduled time is missing")
            if body.skip_existing and any(f"po:{key}" in existing_keys for key in po_keys):
                skipped += 1
                results.append({"row": row_id, "status": "Skipped", "message": "Already appears to exist in DMS.", "reference": row.reference, "po": row.po})
                continue

            if reference_key:
                if reference_key not in truck_number_by_reference:
                    truck_number_by_reference[reference_key] = next_truck_number
                    next_truck_number += 1
                truck_number = truck_number_by_reference[reference_key]
            else:
                truck_number = next_truck_number
                next_truck_number += 1

            payload = _dms_schedule_insert_payload(row, session, business_date, truck_number, upload_model, use_oks_rules)
            if body.dry_run:
                skipped += 1
                results.append({"row": row_id, "status": "Ready", "message": "Ready to upload.", "reference": row.reference, "po": row.po, "truck_number": truck_number})
                continue

            result = _dms_json_request("api/load/insertLoadDetails", payload, session["config"])
            if not _dms_success(result):
                failed += 1
                results.append({
                    "row": row_id,
                    "status": "Failed",
                    "message": str((result or {}).get("userMessage") or (result or {}).get("error") or "DMS rejected the row."),
                    "reference": row.reference,
                    "po": row.po,
                    "truck_number": truck_number,
                })
                continue
            inserted += 1
            for key in po_keys:
                existing_keys.add(f"po:{key}")
            results.append({"row": row_id, "status": "Inserted", "message": "Created in DMS.", "reference": row.reference, "po": row.po, "truck_number": truck_number})
        except Exception as exc:
            failed += 1
            results.append({"row": row_id, "status": "Failed", "message": str(exc), "reference": row.reference, "po": row.po, "truck_number": truck_number})

    return {
        "ok": failed == 0,
        "business_date": business_date,
        "inserted": inserted,
        "skipped": skipped,
        "failed": failed,
        "model_error": model_error,
        "results": results,
    }


@app.get("/api/dms/schedule-upload-model")
def dms_schedule_upload_model():
    session = _ensure_dms_session()
    model = _dms_schedule_upload_model(session)
    insert_model = model.get("insert") if isinstance(model, dict) else None
    buck = session.get("buck") if isinstance(session.get("buck"), dict) else {}
    return {
        "ok": not bool(isinstance(model, dict) and model.get("_error")),
        "model_error": model.get("_error") if isinstance(model, dict) else "DMS upload model was not returned.",
        "insert_fields": sorted(insert_model.keys()) if isinstance(insert_model, dict) else [],
        "areas": buck.get("areas") if isinstance(buck, dict) else [],
    }


@app.post("/api/dms/schedule-upload")
def dms_schedule_upload(body: DmsScheduleUploadIn):
    session = _ensure_dms_session()
    return _run_dms_schedule_upload(session, body, use_oks_rules=True)


@app.get("/api/dms/mn/schedule-upload-model")
def dms_mn_schedule_upload_model():
    session = _ensure_dms_mn_session()
    model = _dms_schedule_upload_model(session)
    insert_model = model.get("insert") if isinstance(model, dict) else None
    buck = session.get("buck") if isinstance(session.get("buck"), dict) else {}
    return {
        "ok": not bool(isinstance(model, dict) and model.get("_error")),
        "model_error": model.get("_error") if isinstance(model, dict) else "DMS upload model was not returned.",
        "insert_fields": sorted(insert_model.keys()) if isinstance(insert_model, dict) else [],
        "areas": buck.get("areas") if isinstance(buck, dict) else [],
    }


@app.post("/api/dms/mn/schedule-upload")
def dms_mn_schedule_upload(body: DmsScheduleUploadIn):
    session = _ensure_dms_mn_session()
    return _run_dms_schedule_upload(session, body, use_oks_rules=False)


# ── PowerView-style billing audit ──────────────────────────────────────────────
# Ported from Orbit's /api/audit/billing (orbit/app.py). One OPS invoice XLSX is
# uploaded; the DMS ranged report is pulled live (no second manual file), then
# cross-referenced for missing/mis-dated/mis-counted pallet breakdowns and rate
# mismatches (Straight Pull tiers, breakdown $/pallet, restack $/pallet).
#
# Rates below come from Orbit's own source (dms_settings.py DEFAULTS for
# ALDIFMN/Minnesota; the ALDIOKS SEED_BILL_CODES table in orbit/app.py for OKS).
# ALDIOKS's Internal Breakdown bill code has no cap in that table, so bd_max_charge
# is left uncapped (0) for OKS -- confirm with the real OPS rate agreement if that's
# wrong.
POWERVIEW_RATES: Dict[str, Dict[str, Any]] = {
    "oks": {
        "bd_rate_per_pallet": 5.0,
        "bd_max_charge": 0.0,
        "restack_rate_per_pallet": 5.0,
        "sp_tiers": {
            "8":    {"min_pal": 1,  "max_pal": 12,   "rate": 60.0},
            "8.01": {"min_pal": 13, "max_pal": 24,   "rate": 95.0},
            "8.02": {"min_pal": 25, "max_pal": 48,   "rate": 105.0},
            "8.03": {"min_pal": 49, "max_pal": None, "rate": 115.0},
        },
    },
    "mn": {
        "bd_rate_per_pallet": 6.0,
        "bd_max_charge": 150.0,
        "restack_rate_per_pallet": 25.0,
        "sp_tiers": {
            "8":    {"min_pal": 1,  "max_pal": 12,   "rate": 42.0},
            "8.01": {"min_pal": 13, "max_pal": 24,   "rate": 85.0},
            "8.02": {"min_pal": 25, "max_pal": 60,   "rate": 92.0},
        },
    },
}

_PV_COL_MAP: Dict[str, List[str]] = {
    "date":       ["Entry Date", "Date", "EntryDate"],
    "po":         ["PONumber", "PO Number", "PO #", "PO"],
    "load_type":  ["Load Type", "LoadType"],
    "desc":       ["Long Description", "Description", "LongDescription"],
    "bill_to":    ["Bill To", "BillTo"],
    "task_qty":   ["Task Qty", "TaskQty"],
    "init_pal":   ["Init Pallets", "InitPallets"],
    "pal_bd":     ["Pallet Brk Down", "Pallet Breakdown", "PalletBrkDown"],
    "carrier":    ["Carrier"],
    "vendor":     ["Vendor"],
    "eclipse_mgr":["Eclipse MGR", "EclipseMgr", "Eclipse Mgr", "Manager", "Mgr"],
    "revenue":    ["Revenue Totals", "Revenue", "Amount"],
    "backhaul":   ["Backhaul"],
    "csh":        ["Csh Chk Card", "CshChkCard", "Cash Check Card"],
}

_PV_BD_RE = re.compile(r'\b(\d+)\s*[Cc](\$)?')


def _pv_fcol(headers: List[str], aliases: List[str]) -> Optional[str]:
    hl = {h.lower().strip(): h for h in headers}
    for a in aliases:
        if a.lower() in hl:
            return hl[a.lower()]
    for a in aliases:
        al = a.lower()
        for kl, korig in hl.items():
            if al in kl or kl in al:
                return korig
    return None


def _pv_parse_date(v: Any) -> Optional[Any]:
    if v is None:
        return None
    if hasattr(v, "date") and callable(getattr(v, "date")):
        return v.date()
    if hasattr(v, "year"):
        return v
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _pv_parse_dms_date(s: Any) -> Optional[Any]:
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _pv_norm_po(po: Any) -> str:
    return re.sub(r'[^0-9]', '', str(po or ''))[:10]


def _pv_tier_for_pallets(tiers: Dict[str, Dict[str, Any]], pallets: float) -> Dict[str, Any]:
    """Pick the tier card whose [min_pal, max_pal] contains `pallets`; falls back
    to the highest tier if `pallets` is above every bounded tier's max."""
    ordered = sorted(tiers.values(), key=lambda t: t["min_pal"])
    for card in ordered:
        if pallets >= card["min_pal"] and (card["max_pal"] is None or pallets <= card["max_pal"]):
            return card
    return ordered[-1] if ordered else {"min_pal": 0, "max_pal": None, "rate": 0.0}


# ── Billing audit export + run history (ported from Orbit's PowerView) ────────
# _billing_audit_tables/_xlsx_bytes/_csv_zip_bytes and the diff helpers below
# are near-verbatim ports of orbit/app.py's billing-audit export and
# "what changed since last audit" comparison, adapted to this app's site
# labels (string "oks"/"mn" instead of a numeric DMS locid).

def _billing_audit_tables(payload: Dict[str, Any]) -> List[Any]:
    """Return [(title, headers, rows, money_col_indices)] from an audit payload."""
    wk = payload.get("week", {}) or {}
    smy = payload.get("summary", {}) or {}
    rec = payload.get("breakdown_recon", {}) or {}
    rng = payload.get("ops_date_range", {}) or {}
    client = payload.get("client_label") or payload.get("dms_location") or "ALDI"
    tables = []

    dms_state = payload.get("dms_status", "")
    if payload.get("dms_error"):
        dms_state = f"{dms_state} ({payload['dms_error']})"
    summary_rows = [
        ["Revenue Week", wk.get("label", "")],
        ["DMS Site", payload.get("dms_location", "")],
        ["OPS Date Range", f"{rng.get('min', '')} - {rng.get('max', '')}"],
        ["DMS Status", dms_state],
        ["", ""],
        ["Total Revenue", smy.get("total_revenue", 0)],
        ["FSPAY Revenue", smy.get("fspay_revenue", 0)],
        [f"{client} Revenue", smy.get("aldi_revenue", 0)],
        ["Total Loads", smy.get("total_loads", 0)],
        ["Total Task Qty", smy.get("total_task_qty", 0)],
        ["Billing Flags", smy.get("flag_count", 0)],
        ["", ""],
        ["DMS c$ Trucks", rec.get("dms_bd_total", 0)],
        ["OPS BD POs", rec.get("ops_bd_total", 0)],
        ["Matched - Correct Date", rec.get("matched_correct", 0)],
        ["Matched - Wrong Date", rec.get("matched_wrong_date", 0)],
        ["Truly Missing", rec.get("truly_missing", 0)],
        ["Est. Unbilled Revenue", rec.get("est_missing_revenue", 0)],
        ["OPS - No DMS c$", rec.get("ops_no_dms", 0)],
        ["Qty Mismatch (revenue)", rec.get("count_mismatch", 0)],
        ["Qty Mismatch Rev Impact", rec.get("count_mismatch_revenue_delta", 0)],
        ["Data Integrity Warnings", rec.get("integrity_warning", 0)],
    ]
    tables.append(("Summary", ["Metric", "Value"], summary_rows, []))

    act_rows = [[
        a.get("load_type", ""), a.get("description", ""), a.get("po_count", 0),
        a.get("task_qty", 0), a.get("total_revenue", 0), a.get("fspay_revenue", 0),
        a.get("aldi_revenue", 0), a.get("avg_rev_po", 0), a.get("rev_pct", 0),
    ] for a in payload.get("activity_breakdown", [])]
    tables.append(("Activity Breakdown",
        ["Load Type", "Description", "PO Count", "Task Qty", "Total Revenue",
         "FSPAY Revenue", f"{client} Revenue", "Avg Rev/PO", "Rev %"],
        act_rows, [5, 6, 7, 8]))

    days = wk.get("days", []) or []
    labels = wk.get("day_labels", []) or []
    dc = payload.get("daily_counts", {}) or {}
    dr = payload.get("daily_revenue", {}) or {}
    daily_rows = []
    for i, dk in enumerate(days):
        counts = dc.get(dk, {}) or {}
        revs = dr.get(dk, {}) or {}
        acts = sorted(set(list(counts.keys()) + list(revs.keys())))
        lbl = labels[i] if i < len(labels) else ""
        for act in acts:
            rv = revs.get(act, {}) or {}
            daily_rows.append([f"{lbl} {dk}".strip(), act, counts.get(act, 0),
                               rv.get("fspay", 0), rv.get("aldi", 0)])
    tables.append(("Daily Load & Revenue",
        ["Day", "Activity", "Loads", "FSPAY Revenue", f"{client} Revenue"],
        daily_rows, [4, 5]))

    bd_rows = []
    for e in rec.get("wrong_date_detail", []):
        bd_rows.append(["Wrong Date", e.get("dms_date", ""), e.get("ops_date", ""),
            e.get("po", ""), e.get("supplier", ""), e.get("carrier", ""),
            e.get("dms_bd", ""), e.get("ops_bd", e.get("ops_task_qty", "")),
            e.get("day_delta", ""), e.get("revenue", ""), e.get("eclipse_mgr", ""),
            e.get("comment", "")])
    for e in rec.get("missing_detail", []):
        bd_rows.append(["Truly Missing", e.get("dms_date", ""), "", e.get("po", ""),
            e.get("supplier", ""), e.get("carrier", ""), e.get("dms_bd", ""), "",
            "", e.get("est_rev", ""), "", e.get("comment", "")])
    for e in rec.get("no_billing_detail", []):
        bd_rows.append(["No Billing Attempt", e.get("dms_date", ""), "", e.get("po", ""),
            e.get("supplier", ""), e.get("carrier", ""), e.get("dms_bd", ""), "",
            "", e.get("est_rev", ""), "", e.get("comment", "")])
    for e in rec.get("count_mismatch_detail", []):
        bd_rows.append(["Qty Mismatch (revenue)", e.get("dms_date", ""), e.get("ops_date", ""),
            e.get("po", ""), e.get("supplier", ""), e.get("carrier", ""),
            e.get("dms_bd", ""), e.get("ops_task_qty", ""), e.get("count_delta", ""),
            e.get("revenue", ""), e.get("eclipse_mgr", ""),
            f"exp ${e.get('expected_rev', 0):g}, delta ${e.get('rev_delta', 0):g} | {e.get('comment', '')}".strip()])
    for e in rec.get("integrity_warning_detail", []):
        bd_rows.append(["Data Integrity (no rev impact)", e.get("dms_date", ""), e.get("ops_date", ""),
            e.get("po", ""), e.get("supplier", ""), e.get("carrier", ""),
            e.get("dms_bd", ""), e.get("ops_bd", ""), e.get("count_delta", ""),
            e.get("revenue", ""), e.get("eclipse_mgr", ""),
            f"ops_bd {e.get('ops_bd','')} vs task_qty {e.get('ops_task_qty','')} | {e.get('comment', '')}".strip()])
    for e in rec.get("ops_no_dms_detail", []):
        bd_rows.append(["Over-billed (OPS no DMS)", "", e.get("ops_date", ""),
            e.get("po", ""), e.get("vendor", ""), e.get("carrier", ""), "",
            e.get("task_qty", ""), "", e.get("revenue", ""), e.get("eclipse_mgr", ""),
            ""])
    tables.append(("Breakdown Detail",
        ["Category", "DMS Date", "OPS Date", "PO", "Supplier/Vendor", "Carrier",
         "DMS c$", "OPS Qty", "Delta", "Revenue/Est", "Eclipse Mgr", "Comment"],
        bd_rows, [10]))

    flag_rows = [[
        f.get("severity", ""), f.get("flag_type", ""), f.get("po", ""), f.get("date", ""),
        f.get("bill_to", ""), f.get("eclipse_mgr", ""), f.get("activity", ""),
        f.get("issue", ""), f.get("billed", ""), f.get("expected", ""),
        f.get("delta", ""), f.get("action", ""),
    ] for f in payload.get("billing_flags", [])]
    tables.append(("Billing Flags",
        ["Severity", "Flag", "PO", "Date", "Bill To", "Eclipse Mgr", "Activity",
         "Issue", "Billed", "Expected", "Delta", "Action"],
        flag_rows, [9, 10, 11]))

    board_rows = [[
        m.get("manager", ""), m.get("total", 0), m.get("errors", 0), m.get("warnings", 0),
        m.get("qty_mismatch", 0), m.get("over_billed", 0), m.get("exposure", 0),
    ] for m in payload.get("error_leaderboard", [])]
    tables.append(("Error Attribution",
        ["Eclipse Manager", "Total Issues", "Rate Errors", "Reviews",
         "Qty Mismatch", "Over-billed", "$ Exposure"],
        board_rows, [7]))

    return tables


def _billing_audit_xlsx_bytes(payload: Dict[str, Any]) -> bytes:
    """Render the audit payload to a styled, multi-sheet, filter/sort-friendly
    workbook: branded header band, zebra striping, borders, number/percent
    formats, text wrapping, and per-sheet highlighting."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    BASE_FONT = Font(name="Calibri", size=10)
    BOLD_FONT = Font(name="Calibri", size=10, bold=True)
    ZEBRA_FILL = PatternFill("solid", fgColor="EEF2F7")
    _edge = Side(style="thin", color="D9DEE4")
    BORDER = Border(left=_edge, right=_edge, top=_edge, bottom=_edge)
    MONEY_FMT = '$#,##0.00'
    PCT_FMT = '0.0"%"'
    WRAP_COLS = {"Issue", "Action", "Comment"}
    WRAP_WIDTH = 42

    SEV_STYLES = {
        "error": (PatternFill("solid", fgColor="F8D7DA"), Font(name="Calibri", size=10, bold=True, color="842029")),
        "warn": (PatternFill("solid", fgColor="FFF3CD"), Font(name="Calibri", size=10, bold=True, color="664D03")),
        "info": (PatternFill("solid", fgColor="E2E3E5"), Font(name="Calibri", size=10, color="41464B")),
    }
    CAT_FILLS = {
        "Wrong Date": PatternFill("solid", fgColor="FFF3CD"),
        "Truly Missing": PatternFill("solid", fgColor="F8D7DA"),
        "No Billing Attempt": PatternFill("solid", fgColor="F3D9EC"),
        "Qty Mismatch": PatternFill("solid", fgColor="FFE5D0"),
        "Over-billed (OPS no DMS)": PatternFill("solid", fgColor="D1E7DD"),
    }
    POS_FONT = Font(name="Calibri", size=10, bold=True, color="B02A37")
    TOP_FILL = PatternFill("solid", fgColor="FDE2E4")

    wb = Workbook()
    first = True
    for title, headers, rows, money_cols in _billing_audit_tables(payload):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = title[:31]
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "1F3A5F"
        ncols = len(headers)
        wrap_idx = {i + 1 for i, h in enumerate(headers) if h in WRAP_COLS}
        pct_idx = {i + 1 for i, h in enumerate(headers) if h.strip() == "Rev %"}

        ws.append(headers)
        ws.row_dimensions[1].height = 22
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

        for ri, r in enumerate(rows, start=2):
            ws.append(r)
            blank = all((v is None or v == "") for v in r)
            if blank:
                continue
            zebra = ZEBRA_FILL if (ri % 2 == 0) else None
            for c in range(1, ncols + 1):
                cell = ws.cell(row=ri, column=c)
                cell.font = BASE_FONT
                cell.border = BORDER
                if zebra:
                    cell.fill = zebra
                is_num = isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
                if c in wrap_idx:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal=("right" if is_num else "left"), vertical="center")
                if c in money_cols:
                    cell.number_format = MONEY_FMT
                elif c in pct_idx:
                    cell.number_format = PCT_FMT

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(len(rows) + 1, 1)}"

        for c in range(1, ncols + 1):
            if c in wrap_idx:
                ws.column_dimensions[get_column_letter(c)].width = WRAP_WIDTH
                continue
            widths = [len(str(headers[c - 1]))]
            widths += [len(str(r[c - 1])) for r in rows if c - 1 < len(r) and r[c - 1] not in (None, "")]
            ws.column_dimensions[get_column_letter(c)].width = min(max(max(widths) + 2, 10), 48)

        if title == "Summary":
            for ri in range(2, len(rows) + 2):
                ws.cell(row=ri, column=1).font = BOLD_FONT
        elif title == "Billing Flags":
            sev_c = 1
            delta_c = headers.index("Delta") + 1 if "Delta" in headers else None
            for ri, r in enumerate(rows, start=2):
                sev = str(r[sev_c - 1]).lower()
                if sev in SEV_STYLES:
                    fill, font = SEV_STYLES[sev]
                    cell = ws.cell(row=ri, column=sev_c)
                    cell.fill = fill; cell.font = font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if delta_c:
                    dcell = ws.cell(row=ri, column=delta_c)
                    if isinstance(dcell.value, (int, float)) and dcell.value > 0:
                        dcell.font = POS_FONT
        elif title == "Breakdown Detail":
            for ri, r in enumerate(rows, start=2):
                fill = CAT_FILLS.get(str(r[0]))
                if fill:
                    cell = ws.cell(row=ri, column=1)
                    cell.fill = fill; cell.font = BOLD_FONT
        elif title == "Error Attribution" and rows:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=2, column=c)
                cell.fill = TOP_FILL
                cell.font = BOLD_FONT

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _billing_audit_csv_zip_bytes(payload: Dict[str, Any]) -> bytes:
    """Render the audit payload to a zip of one CSV per table."""
    import csv as _csv
    import io as _io
    import zipfile as _zip
    buf = BytesIO()
    with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED) as zf:
        for idx, (title, headers, rows, _money) in enumerate(_billing_audit_tables(payload), 1):
            sio = _io.StringIO()
            w = _csv.writer(sio)
            w.writerow(headers)
            w.writerows(rows)
            slug = re.sub(r'[^A-Za-z0-9]+', '-', title).strip('-').lower()
            zf.writestr(f"{idx:02d}-{slug}.csv", sio.getvalue())
    return buf.getvalue()


# ── Billing-audit run comparison — "what changed since the last audit" ───────
def _ba_norm_po(po: Any) -> str:
    return re.sub(r'[^0-9]', '', str(po or ''))[:10]


_BA_PROBLEM_LISTS = (
    ("truly_missing", "missing_detail"),
    ("no_billing", "no_billing_detail"),
    ("qty_mismatch", "count_mismatch_detail"),
    ("integrity_warn", "integrity_warning_detail"),
    ("wrong_date", "wrong_date_detail"),
    ("ops_no_dms", "ops_no_dms_detail"),
)


def _ba_state_map(payload: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """normalized PO -> {state, entry} over the problem lists. POs that are
    clean (matched) aren't enumerated in the payload, so absence == clean."""
    rec = (payload or {}).get("breakdown_recon", {}) or {}
    out: Dict[str, Any] = {}
    for state, key in _BA_PROBLEM_LISTS:
        for e in (rec.get(key) or []):
            po = _ba_norm_po(e.get("po", ""))
            if po and po not in out:
                out[po] = {"state": state, "entry": e}
    return out


def _ba_entry_fields(e: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "supplier": e.get("supplier") or e.get("vendor") or "",
        "carrier": e.get("carrier", ""),
        "dms_bd": e.get("dms_bd"),
        "ops_bd": e.get("ops_bd", e.get("ops_task_qty")),
        "est_rev": e.get("est_rev"),
        "revenue": e.get("revenue"),
        "dms_date": e.get("dms_date", ""),
        "ops_date": e.get("ops_date", ""),
        "eclipse_mgr": e.get("eclipse_mgr", ""),
        "comment": e.get("comment", ""),
    }


def _ba_sig(e: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "dms": (e.get("dms_bd"), e.get("dms_date", ""), e.get("comment", "")),
        "ops": (e.get("ops_bd", e.get("ops_task_qty")), e.get("ops_date", ""), e.get("revenue")),
    }


def _billing_audit_diff(old_payload: Optional[Dict[str, Any]], new_payload: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    old_map, new_map = _ba_state_map(old_payload), _ba_state_map(new_payload)
    resolved, new_issues, still_open, changed = [], [], [], []
    for po in (set(old_map) | set(new_map)):
        o, n = old_map.get(po), new_map.get(po)
        if o and not n:
            resolved.append({"po": po, "was": o["state"], **_ba_entry_fields(o["entry"])})
        elif n and not o:
            new_issues.append({"po": po, "now": n["state"], **_ba_entry_fields(n["entry"])})
        else:
            os_, ns_ = _ba_sig(o["entry"]), _ba_sig(n["entry"])
            if o["state"] == n["state"] and os_ == ns_:
                still_open.append({"po": po, "state": n["state"], **_ba_entry_fields(n["entry"])})
            else:
                changed.append({
                    "po": po, "old_state": o["state"], "new_state": n["state"],
                    "ops_changed": os_["ops"] != ns_["ops"],
                    "dms_changed": os_["dms"] != ns_["dms"],
                    "old": _ba_entry_fields(o["entry"]), "new": _ba_entry_fields(n["entry"]),
                })

    def _flagkey(f: Dict[str, Any]) -> Any:
        return (_ba_norm_po(f.get("po", "")), f.get("flag_type", ""))
    old_flags = {_flagkey(f): f for f in ((old_payload or {}).get("billing_flags") or [])}
    new_flags = {_flagkey(f): f for f in ((new_payload or {}).get("billing_flags") or [])}
    flags_resolved = [old_flags[k] for k in old_flags if k not in new_flags]
    flags_new = [new_flags[k] for k in new_flags if k not in old_flags]

    orec = (old_payload or {}).get("breakdown_recon", {}) or {}
    nrec = (new_payload or {}).get("breakdown_recon", {}) or {}
    osum = (old_payload or {}).get("summary", {}) or {}
    nsum = (new_payload or {}).get("summary", {}) or {}

    def _agg(o: Any, n: Any) -> Dict[str, float]:
        o, n = (o or 0), (n or 0)
        return {"old": round(o, 2), "new": round(n, 2), "delta": round(n - o, 2)}
    aggregate = {
        "est_unbilled": _agg((orec.get("est_missing_revenue", 0) or 0) + (orec.get("est_no_billing_revenue", 0) or 0),
                             (nrec.get("est_missing_revenue", 0) or 0) + (nrec.get("est_no_billing_revenue", 0) or 0)),
        "truly_missing": _agg(orec.get("truly_missing", 0), nrec.get("truly_missing", 0)),
        "no_billing_attempt": _agg(orec.get("no_billing_attempt", 0), nrec.get("no_billing_attempt", 0)),
        "qty_mismatch": _agg(orec.get("count_mismatch", 0), nrec.get("count_mismatch", 0)),
        "qty_mismatch_rev": _agg(orec.get("count_mismatch_revenue_delta", 0), nrec.get("count_mismatch_revenue_delta", 0)),
        "integrity_warn": _agg(orec.get("integrity_warning", 0), nrec.get("integrity_warning", 0)),
        "ops_no_dms": _agg(orec.get("ops_no_dms", 0), nrec.get("ops_no_dms", 0)),
        "matched_correct": _agg(orec.get("matched_correct", 0), nrec.get("matched_correct", 0)),
        "flags": _agg(osum.get("flag_count", 0), nsum.get("flag_count", 0)),
        "total_revenue": _agg(osum.get("total_revenue", 0), nsum.get("total_revenue", 0)),
    }
    return {
        "aggregate": aggregate,
        "resolved": sorted(resolved, key=lambda x: x["po"]),
        "new_issues": sorted(new_issues, key=lambda x: x["po"]),
        "changed": sorted(changed, key=lambda x: x["po"]),
        "still_open": sorted(still_open, key=lambda x: x["po"]),
        "flags_resolved": flags_resolved,
        "flags_new": flags_new,
        "counts": {"resolved": len(resolved), "new_issues": len(new_issues),
                   "changed": len(changed), "still_open": len(still_open),
                   "flags_resolved": len(flags_resolved), "flags_new": len(flags_new)},
    }


def _archive_billing_audit_run(site: str, week_start_str: str, payload: Dict[str, Any],
                                ops_filename: str, ops_bytes: bytes, username: str) -> None:
    """Auto-save this run (result JSON only — DMS is a live, non-reproducible
    pull, so the computed result is the frozen snapshot to diff against later).
    Skips exact-duplicate results (re-run with no change) to avoid clutter."""
    try:
        result_text = json.dumps(payload, sort_keys=True, default=str)
        result_sha = hashlib.sha256(result_text.encode("utf-8")).hexdigest()
        ops_sha = hashlib.sha256(ops_bytes).hexdigest() if ops_bytes else None
        conn = get_db()
        try:
            prev = conn.execute(
                "SELECT result_sha256 FROM billing_audit_runs "
                "WHERE site=? AND week_start=? AND deleted_at IS NULL "
                "ORDER BY run_at DESC, id DESC LIMIT 1",
                (site, week_start_str),
            ).fetchone()
            if not prev or prev["result_sha256"] != result_sha:
                conn.execute(
                    "INSERT INTO billing_audit_runs "
                    "(site, week_start, created_by_user, ops_filename, ops_sha256, result_sha256, result_json) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (site, week_start_str, username, ops_filename, ops_sha, result_sha, result_text),
                )
                conn.commit()
        finally:
            conn.close()
    except Exception as exc:
        print(f"billing audit archive error: {exc}")


def _run_powerview_billing_audit(site: str, ops_bytes: bytes, week_start_str: str) -> Dict[str, Any]:
    rates = POWERVIEW_RATES[site]
    bd_rate = rates["bd_rate_per_pallet"]
    bd_max = rates["bd_max_charge"]
    rs_rate = rates["restack_rate_per_pallet"]
    sp_tiers = rates["sp_tiers"]

    try:
        ws_date = datetime.strptime(week_start_str, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid week_start (expected YYYY-MM-DD).")
    we_date = ws_date + timedelta(days=6)

    # ── Parse OPS XLSX ─────────────────────────────────────────────────────
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(ops_bytes), data_only=True, read_only=True)
        ws_xl = wb.active
        all_xl = list(ws_xl.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse OPS file: {exc}")
    if not all_xl:
        raise HTTPException(status_code=400, detail="OPS file is empty.")
    hdrs = [str(c or "").strip() for c in all_xl[0]]
    ops_rows_raw = [
        {hdrs[i]: (cell if cell is not None else "") for i, cell in enumerate(row) if i < len(hdrs)}
        for row in all_xl[1:]
        if any(cell is not None for cell in row)
    ]

    cols = {k: _pv_fcol(hdrs, v) for k, v in _PV_COL_MAP.items()}

    def rv(row: Dict[str, Any], key: str, default: float = 0) -> float:
        c = cols.get(key)
        if not c:
            return default
        v = row.get(c, default)
        if v is None or v == "":
            return default
        try:
            return float(v)
        except Exception:
            return default

    def sv(row: Dict[str, Any], key: str) -> str:
        c = cols.get(key)
        if not c:
            return ""
        return str(row.get(c, "") or "").strip()

    def ops_rev(row: Dict[str, Any]) -> float:
        r = rv(row, "revenue")
        if r:
            return r
        return rv(row, "backhaul") + rv(row, "csh")

    # ── Split OPS rows into week vs wider window ──────────────────────────
    week_rows: List[Any] = []
    all_dated: List[Any] = []
    for r in ops_rows_raw:
        d = _pv_parse_date(sv(r, "date") or r.get(cols.get("date") or "", ""))
        if d is None:
            continue
        all_dated.append((d, r))
        if ws_date <= d <= we_date:
            week_rows.append((d, r))

    # ── Activity breakdown (target week) ───────────────────────────────────
    act_stats: Dict[Any, Dict[str, float]] = defaultdict(lambda: {"total": 0.0, "fspay": 0.0, "aldi": 0.0, "pos": 0, "tq": 0.0})
    for d, r in week_rows:
        lt = sv(r, "load_type")
        desc = sv(r, "desc")
        bt = sv(r, "bill_to").upper()
        rev = ops_rev(r)
        tq = rv(r, "task_qty")
        key = (lt, desc)
        act_stats[key]["total"] += rev
        act_stats[key]["pos"] += 1
        act_stats[key]["tq"] += tq
        if "FSPAY" in bt:
            act_stats[key]["fspay"] += rev
        else:
            act_stats[key]["aldi"] += rev

    total_rev = sum(v["total"] for v in act_stats.values())
    fspay_rev = sum(v["fspay"] for v in act_stats.values())
    aldi_rev = sum(v["aldi"] for v in act_stats.values())
    total_loads = sum(v["pos"] for v in act_stats.values())
    total_tq = sum(v["tq"] for v in act_stats.values())

    act_out = []
    for (lt_raw, desc), v in sorted(act_stats.items(), key=lambda x: -x[1]["total"]):
        pos = v["pos"]
        tot = v["total"]
        act_out.append({
            "load_type": lt_raw,
            "description": desc,
            "total_revenue": round(tot, 2),
            "fspay_revenue": round(v["fspay"], 2),
            "aldi_revenue": round(v["aldi"], 2),
            "po_count": pos,
            "task_qty": round(v["tq"], 1),
            "avg_rev_po": round(tot / pos, 2) if pos else 0,
            "rev_pct": round(tot / total_rev * 100, 1) if total_rev else 0,
        })

    # ── Daily load count + revenue ─────────────────────────────────────────
    day_keys = [(ws_date + timedelta(days=i)).strftime("%m/%d") for i in range(7)]
    day_labels = ["Sat", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri"]
    daily_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    daily_rev: Dict[str, Dict[str, Dict[str, float]]] = defaultdict(lambda: defaultdict(lambda: {"fspay": 0.0, "aldi": 0.0}))
    for d, r in week_rows:
        dkey = d.strftime("%m/%d")
        desc = sv(r, "desc") or sv(r, "load_type")
        bt = sv(r, "bill_to").upper()
        rev = ops_rev(r)
        daily_counts[dkey][desc] += 1
        if "FSPAY" in bt:
            daily_rev[dkey][desc]["fspay"] += rev
        else:
            daily_rev[dkey][desc]["aldi"] += rev

    # ── Breakdown detection ─────────────────────────────────────────────────
    def has_bd(comment: Any) -> bool:
        return bool(_PV_BD_RE.search(str(comment or '')))

    def bd_cnt(comment: Any) -> int:
        m = _PV_BD_RE.search(str(comment or ''))
        return int(m.group(1)) if m else 0

    def bd_billed(comment: Any) -> bool:
        m = _PV_BD_RE.search(str(comment or ''))
        return bool(m and m.group(2))

    def lt_key(r: Dict[str, Any]) -> str:
        s = sv(r, "load_type")
        try:
            f = float(s)
        except (ValueError, TypeError):
            return s
        return str(int(f)) if f == int(f) else ("%g" % f)

    def is_bd_row(r: Dict[str, Any]) -> bool:
        if "pallet breakdown" in sv(r, "desc").lower():
            return True
        return lt_key(r) == "70"

    # ── Fetch DMS range data live ───────────────────────────────────────────
    dms_rows_ba: List[Dict[str, Any]] = []
    dms_err: Optional[str] = None
    dms_loc_label = ""
    for attempt, force_session in enumerate((False, True)):
        try:
            if site == "oks":
                session = _ensure_dms_session(force=force_session)
            else:
                session = _ensure_dms_mn_session(force=force_session)
            config = session["config"]
            loc = session.get("loc") or {}
            dms_loc_label = loc.get("location") or loc.get("locName") or loc.get("name") or ""
            dms_resp = _dms_post("api/ranged/getRanged", {
                "start": f"{ws_date.month}/{ws_date.day}/{ws_date.year}",
                "end":   f"{we_date.month}/{we_date.day}/{we_date.year}",
                "loc":   loc,
            }, config)
            dms_rows_ba = dms_resp.get("rows", []) or []
            dms_err = None
            break
        except HTTPException as exc:
            dms_err = str(exc.detail)
        except Exception as exc:
            dms_err = str(exc)
        # DMS occasionally answers a login with an empty/null 200 that only
        # shows up as garbage further downstream; retry once with a forced
        # fresh login before giving up.

    # ── Breakdown reconciliation ────────────────────────────────────────────
    dms_bd = [r for r in dms_rows_ba if has_bd(r.get("comments", ""))]
    dms_by_po: Dict[str, Dict[str, Any]] = {}
    for r in dms_bd:
        po = _pv_norm_po(r.get("ponum", ""))
        if po:
            dms_by_po[po] = r

    ops_bd_all = [(d, r) for d, r in all_dated if is_bd_row(r)]
    ops_by_po: Dict[str, Any] = {}
    for d, r in ops_bd_all:
        po = _pv_norm_po(sv(r, "po"))
        if po:
            ops_by_po[po] = (d, r)
    ops_bd_week_pos = {
        _pv_norm_po(sv(r, "po"))
        for d, r in ops_bd_all
        if ws_date <= d <= we_date and _pv_norm_po(sv(r, "po"))
    }

    matched_ok: List[Dict[str, Any]] = []
    matched_wd: List[Dict[str, Any]] = []
    truly_miss: List[Dict[str, Any]] = []
    no_billing: List[Dict[str, Any]] = []
    ops_no_dms: List[Dict[str, Any]] = []

    for dpo, dr in dms_by_po.items():
        dd_date = _pv_parse_dms_date(dr.get("bizDate", ""))
        bdc = bd_cnt(dr.get("comments", ""))
        entry = {
            "dms_date": dd_date.strftime("%m/%d/%Y") if dd_date else "",
            "po": dr.get("ponum", ""),
            "supplier": dr.get("sup", ""),
            "carrier": dr.get("carr", ""),
            "dms_bd": bdc,
            "comment": dr.get("comments", ""),
        }
        if dpo in ops_by_po:
            od, orow = ops_by_po[dpo]
            ops_tq = int(rv(orow, "task_qty"))
            ops_pbd = int(rv(orow, "pal_bd") or ops_tq)
            entry["ops_date"] = od.strftime("%m/%d/%Y") if od else ""
            entry["ops_task_qty"] = ops_tq
            entry["ops_bd"] = ops_pbd
            entry["eclipse_mgr"] = sv(orow, "eclipse_mgr")
            entry["qty_match"] = (ops_tq == bdc) if (ops_tq and bdc) else None
            entry["integrity_match"] = ((ops_pbd == ops_tq == bdc) if (ops_pbd and ops_tq and bdc) else None)
            entry["revenue"] = round(ops_rev(orow), 2)
            entry["day_delta"] = (od - dd_date).days if (od and dd_date) else None
            if dd_date and od and dd_date == od:
                matched_ok.append(entry)
            else:
                matched_wd.append(entry)
        else:
            est_rev = min(bdc * bd_rate, bd_max) if (bdc and bd_max) else (bdc * bd_rate if bdc else 0)
            entry.update({
                "plt_count": int(float(dr.get("qty") or 0)),
                "est_rev": est_rev,
                "billed_flag": bd_billed(dr.get("comments", "")),
            })
            if entry["billed_flag"]:
                truly_miss.append(entry)
            else:
                no_billing.append(entry)

    for opo, (od, orow) in ops_by_po.items():
        if opo not in dms_by_po and opo in ops_bd_week_pos:
            ops_no_dms.append({
                "ops_date": od.strftime("%m/%d/%Y") if od else "",
                "po": sv(orow, "po"),
                "vendor": sv(orow, "vendor"),
                "carrier": sv(orow, "carrier"),
                "eclipse_mgr": sv(orow, "eclipse_mgr"),
                "task_qty": int(rv(orow, "task_qty")),
                "revenue": round(ops_rev(orow), 2),
            })

    def _bd_expected(pallets: float) -> float:
        return min(pallets * bd_rate, bd_max) if bd_max else pallets * bd_rate

    count_mismatch = [e for e in (matched_ok + matched_wd) if e.get("qty_match") is False]
    for e in count_mismatch:
        e["count_delta"] = (e.get("ops_task_qty", 0) or 0) - (e.get("dms_bd", 0) or 0)
        exp_rev = _bd_expected(e.get("dms_bd", 0) or 0)
        e["expected_rev"] = round(exp_rev, 2)
        e["rev_delta"] = round(exp_rev - (e.get("revenue", 0) or 0), 2)

    integrity_warn = [e for e in (matched_ok + matched_wd) if e.get("qty_match") is not False and e.get("integrity_match") is False]
    for e in integrity_warn:
        e["count_delta"] = (e.get("ops_bd", 0) or 0) - (e.get("dms_bd", 0) or 0)

    count_mismatch_rev_delta = round(sum(e.get("rev_delta", 0) or 0 for e in count_mismatch), 2)

    # ── Billing flags ────────────────────────────────────────────────────────
    flags: List[Dict[str, Any]] = []
    for d, r in week_rows:
        lt = sv(r, "load_type")
        ltn = lt_key(r)
        rev = ops_rev(r)
        po = sv(r, "po")
        bt = sv(r, "bill_to")
        dsc = sv(r, "desc")
        emgr = sv(r, "eclipse_mgr")
        ds = d.strftime("%m/%d/%Y") if d else ""
        try:
            ip = float(rv(r, "init_pal"))
        except Exception:
            ip = 0
        try:
            tq = float(rv(r, "task_qty"))
        except Exception:
            tq = 0
        try:
            bd = float(rv(r, "pal_bd"))
        except Exception:
            bd = 0

        if rev < 0:
            flags.append({
                "flag_type": "Credit", "po": po, "date": ds, "bill_to": bt, "eclipse_mgr": emgr,
                "activity": dsc or lt,
                "issue": f"Negative revenue ${rev:.2f}. Verify credit matches agreement.",
                "billed": rev, "expected": None, "delta": None,
                "action": "Confirm credit amount matches current rate agreement.",
                "severity": "info",
            })
            continue

        if ltn in sp_tiers and ip > 0:
            card = sp_tiers[ltn]
            if ip < card["min_pal"] or (card["max_pal"] is not None and ip > card["max_pal"]):
                exp_card = _pv_tier_for_pallets(sp_tiers, ip)
                exp_rate = exp_card["rate"]
                if abs(exp_rate - card["rate"]) > 0.01:
                    flags.append({
                        "flag_type": "Wrong Tier", "po": po, "date": ds, "bill_to": bt, "eclipse_mgr": emgr,
                        "activity": dsc or lt,
                        "issue": (f"{int(ip)} pallets coded {dsc} (rate ${card['rate']:.0f}); "
                                  f"pallet count falls in ${exp_rate:.0f} tier."),
                        "billed": rev, "expected": exp_rate, "delta": round(exp_rate - rev, 2),
                        "action": f"Re-invoice at ${exp_rate}. Delta = ${exp_rate - rev:.0f}.",
                        "severity": "error",
                    })
        elif ltn == "70" or "pallet breakdown" in dsc.lower():
            bd_qty = tq if tq > 0 else bd
            if bd_qty > 0:
                exp_rev = _bd_expected(bd_qty)
                if abs(rev - exp_rev) > 0.01:
                    flags.append({
                        "flag_type": "Rate Mismatch", "po": po, "date": ds, "bill_to": bt, "eclipse_mgr": emgr,
                        "activity": dsc or "Pallet Breakdown",
                        "issue": (f"{int(bd_qty)} BD pallets; expected ${exp_rev:.0f} (${bd_rate:g}/pal"
                                  f"{f', cap ${bd_max:g}' if bd_max else ''}). Billed ${rev:.0f}."),
                        "billed": rev, "expected": exp_rev, "delta": round(exp_rev - rev, 2),
                        "action": f"Review billing. Delta = ${abs(exp_rev - rev):.0f}.",
                        "severity": "error" if rev < exp_rev else "warn",
                    })
        elif ltn == "71.04":
            if ip > 0:
                exp_rev = ip * rs_rate
                if abs(rev - exp_rev) > 0.01:
                    flags.append({
                        "flag_type": "Rate Mismatch", "po": po, "date": ds, "bill_to": bt, "eclipse_mgr": emgr,
                        "activity": dsc or "Pallet Restack",
                        "issue": (f"{int(ip)} pallets; expected ${exp_rev:.0f} (${rs_rate:g}/pal). Billed ${rev:.0f}."),
                        "billed": rev, "expected": exp_rev, "delta": round(exp_rev - rev, 2),
                        "action": f"Re-invoice at ${exp_rev:.0f}. Delta = ${abs(exp_rev - rev):.0f}.",
                        "severity": "error",
                    })

    truly_miss.sort(key=lambda x: x.get("dms_date", ""))
    no_billing.sort(key=lambda x: x.get("dms_date", ""))

    ops_dates = [d for d, _ in all_dated if d]
    ops_min = min(ops_dates).strftime("%m/%d/%Y") if ops_dates else "—"
    ops_max = max(ops_dates).strftime("%m/%d/%Y") if ops_dates else "—"

    est_miss_rev = sum(_bd_expected(r.get("dms_bd", 0)) for r in truly_miss)
    est_no_billing_rev = sum(_bd_expected(r.get("dms_bd", 0)) for r in no_billing)

    # ── Error attribution (Eclipse MGR = OPS PO submitter) ──────────────────
    mgr_stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"errors": 0, "warnings": 0, "qty_mismatch": 0, "over_billed": 0, "exposure": 0.0})
    for f in flags:
        m = (f.get("eclipse_mgr") or "").strip() or "(unattributed)"
        if f.get("severity") == "error":
            mgr_stats[m]["errors"] += 1
        elif f.get("severity") in ("warn", "warning"):
            mgr_stats[m]["warnings"] += 1
        if f.get("delta") not in (None, ""):
            try:
                mgr_stats[m]["exposure"] += abs(float(f["delta"]))
            except (TypeError, ValueError):
                pass
    for e in count_mismatch:
        m = (e.get("eclipse_mgr") or "").strip() or "(unattributed)"
        mgr_stats[m]["qty_mismatch"] += 1
    for e in integrity_warn:
        m = (e.get("eclipse_mgr") or "").strip() or "(unattributed)"
        mgr_stats[m]["warnings"] += 1
    for e in ops_no_dms:
        m = (e.get("eclipse_mgr") or "").strip() or "(unattributed)"
        mgr_stats[m]["over_billed"] += 1
    error_leaderboard = sorted(
        ({"manager": m, "errors": s["errors"], "warnings": s["warnings"],
          "qty_mismatch": s["qty_mismatch"], "over_billed": s["over_billed"],
          "total": s["errors"] + s["warnings"] + s["qty_mismatch"] + s["over_billed"],
          "exposure": round(s["exposure"], 2)}
         for m, s in mgr_stats.items()),
        key=lambda x: (-x["total"], -x["exposure"]),
    )
    error_leaderboard = [x for x in error_leaderboard if x["total"] > 0]

    return {
        "week": {
            "start": ws_date.isoformat(), "end": we_date.isoformat(),
            "label": f"{ws_date.strftime('%m/%d/%Y')} – {we_date.strftime('%m/%d/%Y')}",
            "days": day_keys, "day_labels": day_labels,
        },
        "ops_date_range": {"min": ops_min, "max": ops_max},
        "summary": {
            "total_revenue": round(total_rev, 2), "fspay_revenue": round(fspay_rev, 2),
            "aldi_revenue": round(aldi_rev, 2), "total_loads": total_loads,
            "total_task_qty": round(total_tq, 1), "flag_count": len(flags),
        },
        "activity_breakdown": act_out,
        "daily_counts": {k: dict(v) for k, v in daily_counts.items()},
        "daily_revenue": {k: {a: dict(rv2) for a, rv2 in v.items()} for k, v in daily_rev.items()},
        "breakdown_recon": {
            "dms_bd_total": len(dms_bd), "ops_bd_total": len(ops_bd_all),
            "ops_bd_in_week": sum(1 for d, _ in ops_bd_all if ws_date <= d <= we_date),
            "matched_correct": len(matched_ok), "matched_wrong_date": len(matched_wd),
            "truly_missing": len(truly_miss), "no_billing_attempt": len(no_billing),
            "ops_no_dms": len(ops_no_dms), "count_mismatch": len(count_mismatch),
            "integrity_warning": len(integrity_warn),
            "count_mismatch_revenue_delta": count_mismatch_rev_delta,
            "est_missing_revenue": round(est_miss_rev, 2),
            "est_no_billing_revenue": round(est_no_billing_rev, 2),
            "wrong_date_detail": matched_wd, "missing_detail": truly_miss,
            "no_billing_detail": no_billing, "ops_no_dms_detail": ops_no_dms,
            "count_mismatch_detail": count_mismatch, "integrity_warning_detail": integrity_warn,
        },
        "billing_flags": flags,
        "error_leaderboard": error_leaderboard,
        "dms_status": "error" if dms_err else "ok",
        "dms_error": dms_err,
        "dms_location": dms_loc_label or site.upper(),
        "client_label": dms_loc_label or ("ALDIOKS" if site == "oks" else "ALDIFMN"),
        "audit_rates": {
            "bd_rate_per_pallet": bd_rate, "bd_max_charge": bd_max,
            "restack_rate_per_pallet": rs_rate,
            "sp_tiers": sp_tiers,
        },
        "ops_col_detected": {k: v for k, v in cols.items() if v},
        "ops_rows_total": len(ops_rows_raw),
        "ops_rows_in_week": len(week_rows),
        "dms_rows_total": len(dms_rows_ba),
    }


def _billing_audit_response(payload: Dict[str, Any], fmt: Optional[str]) -> Any:
    """JSON payload, or a downloadable xlsx/csv-zip file if `fmt` asks for one."""
    fmt = (fmt or "").lower()
    if fmt not in ("xlsx", "csv"):
        return payload
    site_slug = re.sub(r'[^A-Za-z0-9]+', '', str(payload.get("dms_location") or "site")) or "site"
    stamp = re.sub(r'[^0-9]', '', str((payload.get("week") or {}).get("start") or ""))[:8] or "latest"
    if fmt == "xlsx":
        body = _billing_audit_xlsx_bytes(payload)
        ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        fname = f"BillingAudit_{site_slug}_{stamp}.xlsx"
    else:
        body = _billing_audit_csv_zip_bytes(payload)
        ctype = "application/zip"
        fname = f"BillingAudit_{site_slug}_{stamp}_csv.zip"
    return Response(content=body, media_type=ctype, headers={"Content-Disposition": f'attachment; filename="{fname}"'})


def _billing_audit_run_summary(row: sqlite3.Row) -> Dict[str, Any]:
    try:
        p = json.loads(row["result_json"])
    except (TypeError, ValueError):
        p = {}
    rec = p.get("breakdown_recon", {}) or {}
    smy = p.get("summary", {}) or {}
    return {
        "id": row["id"], "run_at": row["run_at"],
        "ops_filename": row["ops_filename"] or "",
        "is_baseline": bool(row["is_baseline"]),
        "dms_status": p.get("dms_status"),
        "truly_missing": rec.get("truly_missing", 0),
        "no_billing_attempt": rec.get("no_billing_attempt", 0),
        "qty_mismatch": rec.get("count_mismatch", 0),
        "matched_correct": rec.get("matched_correct", 0),
        "flag_count": smy.get("flag_count", 0),
        "est_unbilled": round((rec.get("est_missing_revenue", 0) or 0) + (rec.get("est_no_billing_revenue", 0) or 0), 2),
    }


@app.post("/api/dms/billing-audit")
async def dms_billing_audit(
    ops_file: UploadFile = File(...),
    week_start: str = Form(...),
    format: Optional[str] = Form(None),
    username: str = Depends(require_auth),
) -> Any:
    ops_bytes = await ops_file.read()
    payload = _run_powerview_billing_audit("oks", ops_bytes, week_start)
    _archive_billing_audit_run("oks", payload["week"]["start"], payload, ops_file.filename or "", ops_bytes, username)
    return _billing_audit_response(payload, format)


@app.post("/api/dms/mn/billing-audit")
async def dms_mn_billing_audit(
    ops_file: UploadFile = File(...),
    week_start: str = Form(...),
    format: Optional[str] = Form(None),
    username: str = Depends(require_auth),
) -> Any:
    ops_bytes = await ops_file.read()
    payload = _run_powerview_billing_audit("mn", ops_bytes, week_start)
    _archive_billing_audit_run("mn", payload["week"]["start"], payload, ops_file.filename or "", ops_bytes, username)
    return _billing_audit_response(payload, format)


class BillingAuditExportIn(BaseModel):
    format: str
    payload: Dict[str, Any]


@app.post("/api/dms/billing-audit/export")
def dms_billing_audit_export(body: BillingAuditExportIn, _: str = Depends(require_auth)) -> Any:
    fmt = (body.format or "").lower()
    if fmt not in ("xlsx", "csv"):
        raise HTTPException(status_code=400, detail="format must be 'xlsx' or 'csv'.")
    if "breakdown_recon" not in body.payload:
        raise HTTPException(status_code=400, detail="Missing or invalid audit payload.")
    return _billing_audit_response(body.payload, fmt)


@app.get("/api/dms/billing-audit/runs")
def dms_billing_audit_runs(site: str, week: str, _: str = Depends(require_auth)) -> Dict[str, Any]:
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT id, run_at, ops_filename, is_baseline, result_json FROM billing_audit_runs "
            "WHERE site=? AND week_start=? AND deleted_at IS NULL ORDER BY run_at DESC, id DESC",
            (site, week),
        ).fetchall()
    finally:
        conn.close()
    return {"runs": [_billing_audit_run_summary(r) for r in rows]}


@app.get("/api/dms/billing-audit/diff")
def dms_billing_audit_diff(
    site: Optional[str] = None,
    week: Optional[str] = None,
    from_id: Optional[int] = None,
    to_id: Optional[int] = None,
    _: str = Depends(require_auth),
) -> Dict[str, Any]:
    conn = get_db()
    try:
        if from_id and to_id:
            a = conn.execute("SELECT * FROM billing_audit_runs WHERE id=? AND deleted_at IS NULL", (from_id,)).fetchone()
            b = conn.execute("SELECT * FROM billing_audit_runs WHERE id=? AND deleted_at IS NULL", (to_id,)).fetchone()
            if not a or not b:
                raise HTTPException(status_code=404, detail="Run not found.")
            old_run, new_run = a, b
        else:
            if not site or not week:
                raise HTTPException(status_code=400, detail="site and week (or from_id and to_id) are required.")
            runs = conn.execute(
                "SELECT * FROM billing_audit_runs WHERE site=? AND week_start=? AND deleted_at IS NULL "
                "ORDER BY run_at DESC, id DESC",
                (site, week),
            ).fetchall()
            if not runs:
                return {"comparison": None, "reason": "no runs"}
            new_run = runs[0]
            baseline = next((r for r in runs if r["is_baseline"] and r["id"] != new_run["id"]), None)
            old_run = baseline or (runs[1] if len(runs) > 1 else None)
            if old_run is None:
                return {"comparison": None, "reason": "no prior run", "current": _billing_audit_run_summary(new_run)}
        try:
            diff = _billing_audit_diff(json.loads(old_run["result_json"]), json.loads(new_run["result_json"]))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=500, detail=f"Could not parse stored runs: {exc}")
        return {
            "comparison": diff,
            "from": _billing_audit_run_summary(old_run),
            "to": _billing_audit_run_summary(new_run),
            "baseline_used": bool(old_run["is_baseline"]),
        }
    finally:
        conn.close()


class BillingAuditBaselineIn(BaseModel):
    baseline: bool = True


@app.post("/api/dms/billing-audit/runs/{run_id}/baseline")
def dms_billing_audit_set_baseline(run_id: int, body: BillingAuditBaselineIn, _: str = Depends(require_auth)) -> Dict[str, Any]:
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT site, week_start FROM billing_audit_runs WHERE id=? AND deleted_at IS NULL", (run_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Run not found.")
        conn.execute(
            "UPDATE billing_audit_runs SET is_baseline=0 WHERE site=? AND week_start=?",
            (row["site"], row["week_start"]),
        )
        if body.baseline:
            conn.execute("UPDATE billing_audit_runs SET is_baseline=1 WHERE id=?", (run_id,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "id": run_id, "is_baseline": bool(body.baseline)}


@app.delete("/api/dms/billing-audit/runs/{run_id}")
def dms_billing_audit_delete_run(run_id: int, _: str = Depends(require_auth)) -> Dict[str, Any]:
    conn = get_db()
    try:
        row = conn.execute("SELECT id FROM billing_audit_runs WHERE id=? AND deleted_at IS NULL", (run_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Run not found.")
        now = datetime.now(timezone.utc).isoformat(timespec="seconds")
        conn.execute("UPDATE billing_audit_runs SET deleted_at=?, is_baseline=0 WHERE id=?", (now, run_id))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "id": run_id, "deleted_at": now}


@app.post("/api/dms/stamp")
def dms_stamp(body: DmsStampIn):
    session = _ensure_dms_session()
    stamp_key = STAMP_TYPE_MAP.get(body.stamp_type.lower().replace(" ", ""), body.stamp_type)
    stamp_time = body.stamp_time or datetime.now(timezone.utc).isoformat()
    payload = {
        "loc":      session["loc"],
        "userinfo": session["userinfo"],
        "buck":     session.get("buck") or {},
        "stampType": stamp_key,
        "stampTime": stamp_time,
    }
    if body.load_id:
        payload["loadId"] = body.load_id
    if body.po:
        payload["po"] = body.po
    candidates = [
        "api/stamp/saveStamp",
        "api/stamp/addStamp",
        "api/stamp/createStamp",
        "api/stamp/stampLoad",
    ]
    last_err = None
    for path in candidates:
        try:
            result = _dms_json_request(path, payload, session["config"])
            ok_flag = True
            if isinstance(result, dict):
                ok_flag = result.get("ok") or result.get("success") or result.get("result") or not result.get("error")
            return {"ok": bool(ok_flag), "endpoint": path, "stamp_type": stamp_key, "stamp_time": stamp_time, "response": result}
        except Exception as e:
            last_err = str(e)
    raise HTTPException(status_code=502, detail=f"DMS stamp failed on all known endpoints. Last error: {last_err}. Open DMS in Chrome DevTools (Network tab), stamp a truck manually, and note the POST URL — then set stamp_endpoint in dms_config.json.")

class ShiftExportIn(BaseModel):
    shift_date: str
    filename: str
    truck_count: int
    csv: str
    trucks: List[Dict[str, Any]] = []

@app.post("/api/portal/export", status_code=201)
def save_shift_export(body: ShiftExportIn):
    safe_name = "".join(c for c in body.filename if c.isalnum() or c in "-_.")
    if not safe_name.endswith(".csv"):
        safe_name += ".csv"
    csv_path = EXPORTS_DIR / safe_name
    csv_path.write_text(body.csv, encoding="utf-8")
    meta_path = EXPORTS_DIR / (safe_name[:-4] + ".json")
    meta_path.write_text(json.dumps({
        "shift_date": body.shift_date,
        "filename": safe_name,
        "truck_count": body.truck_count,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "trucks": body.trucks,
    }, indent=2), encoding="utf-8")
    return {"ok": True, "filename": safe_name}

@app.get("/api/portal/exports")
def list_shift_exports():
    exports = []
    for meta_file in sorted(EXPORTS_DIR.glob("*.json"), reverse=True):
        try:
            data = json.loads(meta_file.read_text(encoding="utf-8"))
            exports.append({
                "shift_date": data.get("shift_date", ""),
                "filename": data.get("filename", ""),
                "truck_count": data.get("truck_count", 0),
                "saved_at": data.get("saved_at", ""),
            })
        except Exception:
            pass
    return {"ok": True, "exports": exports}

@app.get("/api/portal/exports/{filename}")
def download_shift_export(filename: str):
    safe_name = "".join(c for c in filename if c.isalnum() or c in "-_.")
    csv_path = EXPORTS_DIR / safe_name
    if not csv_path.exists() or csv_path.suffix != ".csv":
        raise HTTPException(status_code=404, detail="Export not found.")
    return FileResponse(str(csv_path), media_type="text/csv", filename=safe_name)

class VendorLearnTruck(BaseModel):
    supplier: Optional[str] = None
    dock_min: Optional[int] = None
    shift_date: Optional[str] = None
    source: Optional[str] = "Manual"
    ref: Optional[str] = None

class VendorLearnIn(BaseModel):
    trucks: List[VendorLearnTruck] = []

@app.post("/api/portal/learn", status_code=201)
def portal_learn(body: VendorLearnIn):
    if not body.trucks:
        return {"ok": True, "inserted": 0}
    conn = get_db()
    now_str = datetime.now(timezone.utc).isoformat()
    inserted = 0
    for t in body.trucks:
        vendor = (t.supplier or "").strip().upper()
        if not vendor or t.dock_min is None or t.dock_min <= 0:
            continue
        if t.dock_min > 720:
            continue  # ignore implausible values (>12 hrs)
        conn.execute(
            "INSERT INTO vendor_unload_times (vendor, dock_min, shift_date, source, truck_ref, recorded_at) VALUES (?,?,?,?,?,?)",
            (vendor, t.dock_min, t.shift_date or "", t.source or "Manual", t.ref or "", now_str)
        )
        inserted += 1
    conn.commit()
    conn.close()
    return {"ok": True, "inserted": inserted}

@app.get("/api/portal/vendor-stats")
def portal_vendor_stats():
    conn = get_db()
    rows = conn.execute("""
        SELECT vendor, dock_min, recorded_at
        FROM vendor_unload_times
        WHERE dock_min > 0 AND dock_min <= 720
        ORDER BY vendor, recorded_at DESC
    """).fetchall()
    conn.close()
    from collections import defaultdict
    by_vendor = defaultdict(list)
    for r in rows:
        by_vendor[r[0]].append(r[1])
    latest_seen = {}
    for r in rows:
        if r[0] not in latest_seen:
            latest_seen[r[0]] = r[2]
    stats = []
    for vendor, mins in sorted(by_vendor.items()):
        mins_sorted = sorted(mins)
        n = len(mins_sorted)
        avg = round(sum(mins_sorted) / n)
        p75 = mins_sorted[int(n * 0.75)]
        stats.append({
            "vendor": vendor,
            "avg_min": avg,
            "p75_min": p75,
            "min_min": mins_sorted[0],
            "max_min": mins_sorted[-1],
            "count": n,
            "last_seen": latest_seen.get(vendor, ""),
        })
    return {"ok": True, "stats": stats}

@app.get("/")
def serve_app():
    # Never cache the app shell — the whole UI is inline in index.html, so a stale
    # cached copy serves stale JS and can blank the page after a deploy. Force the
    # browser to revalidate every load so users always get the latest code.
    return FileResponse(
        str(BASE_DIR / "index.html"),
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.get("/historical-data.js")
def serve_historical_data():
    return FileResponse(
        str(BASE_DIR / "historical-data.js"),
        media_type="application/javascript",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )

